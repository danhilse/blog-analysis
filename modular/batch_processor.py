import json
import pandas as pd
import os
from typing import Dict, Optional
import openpyxl
from article_processor import (
    clean_content, calculate_word_count, load_yoast_keywords,
    import_performance_data, count_personal_pronouns, format_seo_data, parse_date
)
from depr.ai_analysis import (
    analyze_content_categorization, analyze_seo,
    analyze_tone_voice, analyze_quality_brand_fit, cost_tracker
)
from excel_styler import ExcelStyler
from urllib.parse import urlparse


class BatchProcessor:
    def __init__(self, json_file: str, excel_file: str):
        """
        Initialize the batch processor.

        Args:
            json_file: Path to the JSON file containing all articles
            excel_file: Path to the Excel file for output
        """
        self.json_file = json_file
        self.excel_file = excel_file
        self.url_to_keyword = load_yoast_keywords()
        self.performance_df = import_performance_data()
        self.last_processed_index = None

    def process_batch(self, start_index: int, batch_size: int) -> Optional[int]:
        """
        Process a batch of articles and append to Excel.

        Args:
            start_index: Starting index in the JSON file
            batch_size: Number of articles to process

        Returns:
            Optional[int]: Index of last successfully processed article or None if error
        """
        try:
            # Load all articles from JSON
            with open(self.json_file, 'r') as file:
                all_data = json.load(file)

            # Get the batch of articles
            batch_end = start_index + batch_size
            batch_articles = all_data['analyses']['blog'][start_index:batch_end]

            if not batch_articles:
                print("No articles found in specified range")
                return None

            # Process each article
            processed_articles = []
            for i, article in enumerate(batch_articles):
                current_index = start_index + i
                try:
                    processed_row = self._process_single_article(article)
                    processed_articles.append(processed_row)
                    self.last_processed_index = current_index
                    print(f"Processed article {current_index}")
                except Exception as e:
                    print(f"Error processing article at index {current_index}: {str(e)}")
                    return self.last_processed_index

            # Convert to DataFrame
            df = pd.DataFrame(processed_articles)

            # Append to Excel or create new file
            if not os.path.exists(self.excel_file):
                self._create_new_excel(df)
            else:
                self._append_to_excel(df)

            return self.last_processed_index

        except Exception as e:
            print(f"Batch processing error: {str(e)}")
            return self.last_processed_index

    def _process_single_article(self, article: Dict) -> Dict:
        """Process a single article and return its data."""
        cost_tracker.reset()

        # Extract basic content
        content = article.get('content', '')
        clean_text = clean_content(content)

        # Extract basic info
        basic_info = article.get('basic_info', {})
        url = basic_info.get('url', 'No URL')
        title = basic_info.get('title', 'No Title')

        # Get pronoun analysis
        pronoun_analysis = count_personal_pronouns(content)

        # Get target keyword
        target_keyword = self.url_to_keyword.get(url, 'Not Found')

        # Extract slug for performance data
        slug = urlparse(url).path.strip('/').split('/')[-1] if url != 'No URL' else 'No Slug'

        # Get multimedia assessment
        multimedia = article.get('multimedia_assessment', {})
        header_image = multimedia.get('header_image') or {}

        # Process content and get analyses
        categorization = analyze_content_categorization(clean_text)
        seo_data = format_seo_data(basic_info, article.get('seo_analysis', {}), target_keyword)
        seo_results = analyze_seo(clean_text, seo_data)
        tone_results = analyze_tone_voice(clean_text)
        quality_results = analyze_quality_brand_fit(clean_text)

        # Get performance metrics
        performance_metrics = self._get_performance_metrics(slug)

        # Create the row data
        return {
            'Title': title,
            'URL': url,
            'Slug': slug,
            'Publication Date': parse_date(basic_info.get('publication_date')),
            'Modified Date': parse_date(basic_info.get('modified_date')),
            'Word Count': calculate_word_count(clean_text),
            'Reading Level (Gunning Fog)': round(float(quality_results.get('Reading Level', 0)), 1),

            # Quality & Brand Fit
            'Overall Quality Score': int(quality_results.get('Overall Quality Score', 0)),
            'Topic Relevance': quality_results.get('Topic Relevance', 'Error'),
            'Brand Alignment': quality_results.get('Brand Alignment', 'Needs Work'),
            'Quality Notes/Recommendations': quality_results.get('Quality Notes', ''),
            'Brand Alignment Notes': quality_results.get('Brand Alignment Notes', ''),

            # Add all other fields...
            # [Rest of the fields exactly as in your original processing]

            'API Cost': f"${cost_tracker.cost:.4f}"
        }

    def _get_performance_metrics(self, slug: str) -> Dict:
        """Get performance metrics for a slug."""
        if self.performance_df is not None and slug in self.performance_df.index:
            metrics = self.performance_df.loc[slug]
            return {
                'Total Views': metrics.get('Views', 0),
                'Total Users': metrics.get('Total users', 0),
                'Total Sessions': metrics.get('Sessions', 0),
                'Engagement Rate': metrics.get('Engagement rate', 0.0),
                'Average Time on Page': metrics.get('Average session duration', 0.0),
                'Bounce Rate': metrics.get('Bounce rate', 0.0)
            }
        return {
            'Total Views': 0,
            'Total Users': 0,
            'Total Sessions': 0,
            'Engagement Rate': 0.0,
            'Average Time on Page': 0.0,
            'Bounce Rate': 0.0
        }

    def _create_new_excel(self, df: pd.DataFrame):
        """Create a new Excel file with headers."""
        with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

    def _append_to_excel(self, df: pd.DataFrame):
        """Append data to existing Excel file."""
        # Read existing data
        existing_df = pd.read_excel(self.excel_file)

        # Concatenate with new data
        combined_df = pd.concat([existing_df, df], ignore_index=True)

        # Write back to Excel
        with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
            combined_df.to_excel(writer, index=False, sheet_name='Sheet1')

    def _append_to_excel(self, df: pd.DataFrame):
        """Append data to existing Excel file."""
        book = openpyxl.load_workbook(self.excel_file)
        writer = pd.ExcelWriter(self.excel_file, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        # Get the first worksheet
        ws = writer.sheets['Sheet1']

        # Find the next empty row
        next_row = ws.max_row + 1

        # Write the new data starting from the next empty row
        for idx, row in df.iterrows():
            for col, value in enumerate(row, 1):
                ws.cell(row=next_row + idx, column=col, value=value)

        writer.save()


def style_excel_file(excel_file: str):
    """Apply styling to the complete Excel file."""
    styler = ExcelStyler()
    styler.apply_full_styling(excel_file)