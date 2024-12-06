import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


class ExcelStyler:
    def __init__(self):
        # Define brand colors with lighter shades for subheaders
        self.section_colors = {
            'Title': {
                'header': '#193661',  # Dark blue
                'subheader': '#C1C9D4'  # Lighter blue
            },
            'Basic Information': {
                'header': '#00babe',  # Teal
                'subheader': '#E5F9F9'  # Lighter teal
            },
            'Quality & Brand Fit': {
                'header': '#e34e64',  # Salmon
                'subheader': '#FFEFEF'  # Lighter salmon
            },
            'Tone & Voice': {
                'header': '#193661',  # Dark blue
                'subheader': '#C1C9D4'  # Lighter blue
            },
            'SEO Analysis': {
                'header': '#00babe',  # Teal
                'subheader': '#E5F9F9'  # Lighter teal
            },
            'Multimedia Assessment': {
                'header': '#e34e64',  # Salmon
                'subheader': '#FFEFEF'  # Lighter salmon
            },
            'Content Categorization': {
                'header': '#193661',  # Dark blue
                'subheader': '#C1C9D4'  # Lighter blue
            },
            'Performance Metrics': {
                'header': '#00babe',  # Teal
                'subheader': '#E5F9F9'  # Lighter teal
            },
            'Cost Analysis': {
                'header': '#e34e64',  # Salmon
                'subheader': '#FFEFEF'  # Lighter salmon
            }
        }

        # Define sections and their column ranges
        self.sections = {
            'Title': (1, 1),
            'Basic Information': (2, 7),
            'Quality & Brand Fit': (8, 12),
            'Tone & Voice': (13, 19),
            'SEO Analysis': (20, 29),
            'Multimedia Assessment': (30, 36),
            'Content Categorization': (37, 43),
            'Performance Metrics': (44, 49),
            'Cost Analysis': (50, 50)
        }

        # Define common styles
        self.thin_border = Border(
            left=Side(style='thin', color='E3E3E3'),
            right=Side(style='thin', color='E3E3E3'),
            top=Side(style='thin', color='E3E3E3'),
            bottom=Side(style='thin', color='E3E3E3')
        )

    def apply_full_styling(self, filename: str):
        """Apply all styles to the Excel file."""
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        self._setup_basic_formatting(ws)
        self._add_section_headers(ws)
        self._style_data_cells(ws)
        self._add_conditional_formatting(ws)
        self._add_no_match_highlighting(ws)
        self._apply_column_widths(ws)

        wb.save(filename)

    def _setup_basic_formatting(self, ws):
        """Apply basic worksheet formatting."""
        # Freeze panes
        ws.freeze_panes = 'B3'

        # Style metric headers (second row)
        for col_idx, cell in enumerate(ws[2], 1):
            for section, (start, end) in self.sections.items():
                if start <= col_idx <= end:
                    subheader_color = self.section_colors[section]['subheader'].replace('#', '')
                    cell.fill = PatternFill(start_color=subheader_color,
                                            end_color=subheader_color,
                                            fill_type='solid')
                    cell.font = Font(bold=True, color='444444')
                    cell.alignment = Alignment(horizontal='center',
                                               vertical='center',
                                               wrap_text=True)
                    break

    def _add_section_headers(self, ws):
        """Add and style section headers."""
        ws.insert_rows(1)
        current_col = 1
        for section, (start, end) in self.sections.items():
            if current_col <= end:
                cell = ws.cell(row=1, column=current_col)
                cell.value = section
                cell.font = Font(bold=True, color='FFFFFF')

                header_color = self.section_colors[section]['header'].replace('#', '')
                cell.fill = PatternFill(start_color=header_color,
                                        end_color=header_color,
                                        fill_type='solid')

                cell.alignment = Alignment(horizontal='center',
                                           vertical='center',
                                           wrap_text=True)

                if start != end:
                    ws.merge_cells(start_row=1,
                                   start_column=current_col,
                                   end_row=1,
                                   end_column=end)
                current_col = end + 1

    def _style_data_cells(self, ws):
        """Style all data cells including special URL formatting."""
        # Add alternating row colors
        for row in range(3, ws.max_row + 1):
            fill_color = 'F7F9FB' if row % 2 == 0 else 'FFFFFF'
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)

                # Special formatting for URL column
                if col == 2:  # URL column (B)
                    cell.alignment = Alignment(horizontal='left',
                                               vertical='center',
                                               wrap_text=False)
                    cell.font = Font(color='0563C1', underline='single')
                    if cell.value and isinstance(cell.value, str):
                        cell.hyperlink = cell.value
                else:
                    cell.alignment = Alignment(horizontal='left',
                                               vertical='center',
                                               wrap_text=True)
                    cell.font = Font(color='444444')

                # Apply fill color if no other fill exists
                if not cell.fill.start_color.rgb:
                    cell.fill = PatternFill(start_color=fill_color,
                                            end_color=fill_color,
                                            fill_type='solid')

                cell.border = self.thin_border

    def _add_conditional_formatting(self, ws):
        """Add conditional formatting rules to specific columns."""
        header_row = [cell.value for cell in ws[2]]

        # Score columns for color scale rules
        score_columns = [
            'Overall Quality Score',
            'Natural/Conversational Score',
            'Authentic/Approachable Score',
            'Gender-Neutral/Inclusive Score',
            'Keyword Integration Score',
            'Meta Description Quality Score',
            'Reading Level (Gunning Fog)'
        ]

        for col_name in score_columns:
            if col_name in header_row:
                col_letter = get_column_letter(header_row.index(col_name) + 1)
                ws.conditional_formatting.add(
                    f'{col_letter}3:{col_letter}{ws.max_row}',
                    ColorScaleRule(
                        start_type='num', start_value=0, start_color='FF0000',
                        mid_type='num', mid_value=50, mid_color='FFFF00',
                        end_type='num', end_value=100, end_color='00FF00'
                    )
                )

        # Format cost column
        if 'API Cost' in header_row:
            cost_col = header_row.index('API Cost') + 1
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=cost_col)
                cell.number_format = '$#,##0.00000'

    def _add_no_match_highlighting(self, ws):
        """Add red highlighting for 'No Clear Match' values."""
        header_row = [cell.value for cell in ws[2]]

        target_columns = [
            'Primary Category',
            'Solution Topic',
            'Use Case',
            'Customer Journey Stage',
            'CMO Priority',
            'Marketing Activity Type',
            'Target Audience'
        ]

        highlight_values = [
            'No Clear Match',
            'NONE',
            'No Clear Topic',
            'No Clear Activity Type',
            'No Clear Audience'
        ]

        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True)

        for col_name in target_columns:
            if col_name in header_row:
                col_idx = header_row.index(col_name) + 1
                for row in range(3, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value in highlight_values:
                        cell.fill = red_fill
                        cell.font = white_font

    def _apply_column_widths(self, ws):
        """Set appropriate column widths."""
        # Set URL column width
        ws.column_dimensions['B'].width = 15

        # Set all other columns
        for col in range(1, ws.max_column + 1):
            if get_column_letter(col) != 'B':
                ws.column_dimensions[get_column_letter(col)].width = 30