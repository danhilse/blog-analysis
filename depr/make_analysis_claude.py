import json
from textstat import textstat
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import pandas as pd
from datetime import datetime
from urllib.parse import urlparse
import os
from typing import Dict, List, Tuple
from pathlib import Path



def append_to_excel(df, output_file):
    """Append new rows to existing Excel file while preserving formatting."""

    # Load existing workbook
    book = openpyxl.load_workbook(output_file)
    sheet = book.active

    # Get the next empty row (after headers)
    next_row = sheet.max_row + 1

    # Convert DataFrame to list of lists
    values = df.values.tolist()

    # Add new rows
    for row_index, row_data in enumerate(values, start=next_row):
        for col_index, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)

            # Maintain URL formatting for URL column
            if col_index == 2:  # URL column
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                cell.font = Font(color='0563C1', underline='single')
                if value and isinstance(value, str):
                    cell.hyperlink = value
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.font = Font(color='444444')

    # Apply conditional formatting and styling to new rows
    apply_row_formatting(sheet, next_row)

    # Add the cost column formatting
    cost_col = None
    for idx, cell in enumerate(sheet[2]):  # Check header row
        if cell.value == 'API Cost':
            cost_col = idx + 1
            break

    if cost_col:
        for row in range(next_row, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=cost_col)
            cell.number_format = '$#,##0.00000'

    # Save the workbook
    book.save(output_file)


def apply_row_formatting(sheet, start_row):
    """Apply formatting to newly added rows."""

    for row in range(start_row, sheet.max_row + 1):
        # Alternate row colors
        fill_color = 'F7F9FB' if row % 2 == 0 else 'FFFFFF'

        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if not cell.fill.start_color.rgb:  # Only apply if no other fill color is present
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

            # Add borders
            cell.border = Border(
                left=Side(style='thin', color='E3E3E3'),
                right=Side(style='thin', color='E3E3E3'),
                top=Side(style='thin', color='E3E3E3'),
                bottom=Side(style='thin', color='E3E3E3')
            )


def clean_content(json_input):
    """
    Cleans and converts JSON content to Markdown.

    Parameters:
        json_input (str): JSON string containing the "content" field.

    Returns:
        str: Cleaned Markdown content.
    """
    try:
        # Parse the JSON input
        data = json.loads(json_input)
        content = data.get("content", "")
    except json.JSONDecodeError:
        # If input is not a valid JSON, assume it's raw content
        content = json_input

    # Remove image markers like [CONTENT IMAGE: ...]
    content = re.sub(r'\[CONTENT IMAGE:.*?\]', '', content)

    # Remove source lines like Source: https://...
    content = re.sub(r'Source:\s*https?://\S+', '', content)

    # Convert 'H2:' headers to Markdown '## '
    content = re.sub(r'H2:\s*', '## ', content)

    # Normalize line breaks and remove excessive whitespace
    content = re.sub(r'\n{3,}', '\n\n', content)  # Replace 3+ newlines with 2
    content = re.sub(r'\n\s*\n', '\n\n', content)  # Ensure double newlines for paragraphs
    content = re.sub(r'[ \t]+', ' ', content)      # Replace multiple spaces/tabs with single space

    # Remove any remaining unwanted brackets or markdown artifacts
    content = re.sub(r'\[.*?\]', '', content)

    # Trim leading and trailing whitespace
    content = content.strip()

    return content

def calculate_word_count(content):
    clean_text = clean_content(content)
    return len(clean_text.split())


def add_no_match_highlighting(ws):
    """
    Adds conditional highlighting to mark cells containing 'No Clear Match', 'NONE',
    or similar values in red.

    Parameters:
        ws (openpyxl.worksheet.worksheet.Worksheet): The active worksheet
    """
    # Get header row values
    header_row = [cell.value for cell in ws[2]]  # Headers are in row 2

    # Columns to check for "No Clear" values
    target_columns = [
        'Primary Category',
        'Solution Topic',
        'Use Case',
        'Customer Journey Stage',
        'CMO Priority',
        'Marketing Activity Type',
        'Target Audience'
    ]

    # Get column indices for target columns
    column_indices = {col: header_row.index(col) + 1 for col in target_columns if col in header_row}

    # Values to highlight in red
    highlight_values = [
        'No Clear Match',
        'NONE',
        'No Clear Topic',
        'No Clear Activity Type',
        'No Clear Audience'
    ]

    # Red fill for matching cells
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)

    # Apply highlighting to matching cells
    for row in range(3, ws.max_row + 1):  # Start from data rows
        for col_name, col_idx in column_indices.items():
            cell = ws.cell(row=row, column=col_idx)
            if cell.value in highlight_values:
                cell.fill = red_fill
                cell.font = white_font


def add_conditional_formatting(ws):
    """Add conditional formatting rules to specific columns."""
    # Get header row values and their indices
    header_row = [cell.value for cell in ws[2]]

    # Mapping of columns to their indices
    columns = {
        'Overall Quality Score': header_row.index('Overall Quality Score') + 1,
        'Topic Relevance': header_row.index('Topic Relevance') + 1,
        'Brand Alignment': header_row.index('Brand Alignment') + 1,
        'Natural/Conversational Score': header_row.index('Natural/Conversational Score') + 1,
        'Authentic/Approachable Score': header_row.index('Authentic/Approachable Score') + 1,
        'Gender-Neutral/Inclusive Score': header_row.index('Gender-Neutral/Inclusive Score') + 1,
        'Personal Pronoun Count': header_row.index('Personal Pronoun Count') + 1,
        'Keyword Integration Score': header_row.index('Keyword Integration Score') + 1,
        'Meta Description Quality Score': header_row.index('Meta Description Quality Score') + 1,
        'Outdated Widgets Count': header_row.index('Outdated Widgets Count') + 1,
        'Word Count': header_row.index('Word Count') + 1

    }

    # Define fill patterns
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow_green_fill = PatternFill(start_color='9ACD32', end_color='9ACD32', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    # Add ColorScale rules for score columns
    score_columns = ['Overall Quality Score', 'Natural/Conversational Score',
                     'Authentic/Approachable Score', 'Gender-Neutral/Inclusive Score',
                     'Keyword Integration Score', 'Meta Description Quality Score']

    for col_name in score_columns:
        col_letter = get_column_letter(columns[col_name])
        ws.conditional_formatting.add(
            f'{col_letter}3:{col_letter}{ws.max_row}',
            ColorScaleRule(
                start_type='num', start_value=0, start_color='FF0000',
                mid_type='num', mid_value=50, mid_color='FFFF00',
                end_type='num', end_value=100, end_color='00FF00'
            )
        )

    # Add Word Count formatting rules
    word_count_col = get_column_letter(columns['Word Count'])
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=columns['Word Count'])
        try:
            count = int(cell.value)
            if 1000 <= count <= 1200:
                cell.fill = green_fill
            elif (800 <= count < 1000) or (1201 <= count <= 1400):
                cell.fill = yellow_fill
            else:
                cell.fill = orange_fill
        except (ValueError, TypeError):
            pass


    # Add rules for Topic Relevance
    topic_col = get_column_letter(columns['Topic Relevance'])
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=columns['Topic Relevance'])
        if cell.value == "Tangentially Related":
            cell.fill = yellow_fill
        elif cell.value == "Off Topic":
            cell.fill = red_fill

    # Add rules for Brand Alignment
    brand_col = get_column_letter(columns['Brand Alignment'])
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=columns['Brand Alignment'])
        if cell.value == "On Brand":
            cell.fill = green_fill
        elif cell.value == "Mostly on Brand":
            cell.fill = yellow_green_fill
        elif cell.value == "Needs Work":
            cell.fill = yellow_fill
        elif cell.value == "Not on Brand":
            cell.fill = red_fill

    # Add rules for Outdated Widgets Count
    widgets_col = get_column_letter(columns['Outdated Widgets Count'])
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=columns['Outdated Widgets Count'])
        try:
            if int(cell.value) > 0:
                cell.fill = red_fill
        except (ValueError, TypeError):
            pass

    # Add rules for Personal Pronoun Count
    pronoun_col = get_column_letter(columns['Personal Pronoun Count'])
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=columns['Personal Pronoun Count'])
        try:
            if int(cell.value) > 0:
                cell.fill = red_fill
        except (ValueError, TypeError):
            pass



def style_excel_file(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Extract title from first row and remove it
    title_cell = ws['A2'].value  # Save title value

    # Define brand colors with lighter shades for subheaders
    section_colors = {
        'Title': {
            'header': '#193661',  # Dark blue
            'subheader': '#C1C9D4'  # Lighter blue
        },
        'Basic Information': {
            'header': '#00babe',  # Teal
            'subheader': '#E5F9F9'  # Lighter teal
        },
        'Quality & Brand Fit': {
            'header': '#e34e64',  # Salmon
            'subheader': '#FFEFEF'  # Lighter salmon
        },
        'Tone & Voice': {
            'header': '#193661',  # Dark blue
            'subheader': '#C1C9D4'  # Lighter blue
        },
        'SEO Analysis': {
            'header': '#00babe',  # Teal
            'subheader': '#E5F9F9'  # Lighter teal
        },
        'Multimedia Assessment': {
            'header': '#e34e64',  # Salmon
            'subheader': '#FFEFEF'  # Lighter salmon
        },
        'Content Categorization': {
            'header': '#193661',  # Dark blue
            'subheader': '#C1C9D4'  # Lighter blue
        },
        'Performance Metrics': {
            'header': '#00babe',  # Teal
            'subheader': '#E5F9F9'  # Lighter teal
        },
        'Cost Analysis': {            # Added missing section
            'header': '#e34e64',      # Salmon
            'subheader': '#FFEFEF'    # Lighter salmon
        }
    }

    # Define sections and their ranges
    sections = {
        'Title': (1, 1),
        'Basic Information': (2, 7),  # Updated to include Modified Date
        'Quality & Brand Fit': (8, 12),
        'Tone & Voice': (13, 19),
        'SEO Analysis': (20, 29),  # Updated for split H2/H3
        'Multimedia Assessment': (30, 36),
        'Content Categorization': (37, 43),
        'Performance Metrics': (44, 49),
        'Cost Analysis': (50, 50)  # New section
    }

    # Freeze top two rows and first column
    ws.freeze_panes = 'B3'

    # Set specific column widths
    ws.column_dimensions['B'].width = 15  # URL column
    for col in range(1, ws.max_column + 1):
        if get_column_letter(col) != 'B':  # All non-URL columns
            ws.column_dimensions[get_column_letter(col)].width = 30

    # # Replace the original column width setting code with:
    # TINY = 4  # Fit 4 characters
    # SMALL = 8  # Fit 8 characters
    # HALF = 15  # Half width
    # DEFAULT = 30  # No change
    #
    # width_map = {
    #     0: DEFAULT, 1: DEFAULT, 2: DEFAULT, 3: HALF, 4: HALF, 5: HALF,
    #     6: TINY, 7: TINY, 8: HALF, 9: DEFAULT, 10: DEFAULT, 11: DEFAULT,
    #     12: TINY, 13: TINY, 14: TINY, 15: TINY, 16: TINY, 17: DEFAULT,
    #     18: TINY, 19: DEFAULT, 20: SMALL, 21: TINY, 22: TINY, 23: TINY,
    #     24: TINY, 25: TINY, 26: TINY, 27: DEFAULT, 28: DEFAULT, 29: TINY,
    #     30: SMALL, 31: SMALL, 32: TINY, 33: DEFAULT, 34: DEFAULT, 35: SMALL,
    #     36: TINY, 37: DEFAULT, 38: DEFAULT, 39: DEFAULT, 40: DEFAULT,
    #     41: DEFAULT, 42: DEFAULT, 43: DEFAULT, 44: DEFAULT, 45: TINY,
    #     46: TINY, 47: TINY, 48: TINY, 49: TINY, 50: TINY
    # }
    #
    # for col_idx, width in width_map.items():
    #     column_letter = get_column_letter(col_idx + 1)
    #     ws.column_dimensions[column_letter].width = width

    # Style metric headers (second row)
    row2 = ws[1]  # Get the second row (index 1)
    for col_idx, cell in enumerate(row2, 1):
        # Find which section this column belongs to
        for section, (start, end) in sections.items():
            if start <= col_idx <= end:
                subheader_color = section_colors[section]['subheader'].replace('#', '')
                cell.fill = PatternFill(start_color=subheader_color,
                                      end_color=subheader_color,
                                      fill_type='solid')
                cell.font = Font(bold=True, color='444444')
                cell.alignment = openpyxl.styles.Alignment(horizontal='center',
                                                         vertical='center',
                                                         wrap_text=True)
                break

    # Add and style section headers
    ws.insert_rows(1)
    current_col = 1
    for section, (start, end) in sections.items():
        if current_col <= end:
            cell = ws.cell(row=1, column=current_col)
            cell.value = section
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=section_colors[section]['header'].replace('#', ''),
                                  end_color=section_colors[section]['header'].replace('#', ''),
                                  fill_type='solid')
            cell.alignment = openpyxl.styles.Alignment(horizontal='center',
                                                     vertical='center',
                                                     wrap_text=True)

            if start != end:
                ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=end)
            current_col = end + 1

    # Get column indices for conditional formatting
    header_row = [cell.value for cell in ws[2]]  # Get header row values
    reading_level_col = header_row.index('Reading Level (Gunning Fog)') + 1
    header_image_width_col = header_row.index('Header Image Width') + 1

    # Add conditional formatting for Word Count and Reading Level

    ws.conditional_formatting.add(
        f'{get_column_letter(reading_level_col)}3:{get_column_letter(reading_level_col)}{ws.max_row}',
        ColorScaleRule(
            start_type='min',
            start_color='FF0000',  # Red
            mid_type='percentile',
            mid_value=50,
            mid_color='FFFF00',  # Yellow
            end_type='max',
            end_color='00FF00'  # Green
        )
    )

    # Add conditional formatting for Header Image Width
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=header_image_width_col)
        try:
            width = int(cell.value)
            if width >= 1200:
                cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
            elif width >= 800:
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
            else:
                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
        except (ValueError, TypeError):
            pass

    # Format data cells (including URL special formatting)
    for row in ws.iter_rows(min_row=3):  # Start from data rows
        for cell in row:
            # Special formatting for URL column
            if cell.column == 2:  # URL column (B)
                cell.alignment = openpyxl.styles.Alignment(horizontal='left',
                                                         vertical='center',
                                                         wrap_text=False)
                cell.font = Font(color='0563C1', underline='single')
                if cell.value and isinstance(cell.value, str):
                    cell.hyperlink = cell.value
            else:
                cell.alignment = openpyxl.styles.Alignment(horizontal='left',
                                                         vertical='center',
                                                         wrap_text=True)
                cell.font = Font(color='444444')

    # Add alternating row colors for better readability
    for row in range(3, ws.max_row + 1):
        fill_color = 'F7F9FB' if row % 2 == 0 else 'FFFFFF'
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if not cell.fill.start_color.rgb:  # Only apply if no other fill color is present
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    # Add thin borders to all cells
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin', color='E3E3E3'),
        right=openpyxl.styles.Side(style='thin', color='E3E3E3'),
        top=openpyxl.styles.Side(style='thin', color='E3E3E3'),
        bottom=openpyxl.styles.Side(style='thin', color='E3E3E3')
    )

    # Add the new highlighting for "No Clear Match" values
    add_no_match_highlighting(ws)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    cost_col = header_row.index('API Cost') + 1
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=cost_col)
        cell.number_format = '$#,##0.00000'

    add_conditional_formatting(ws)

    # Restore title in its new position
    ws['A3'] = title_cell

    wb.save(filename)

def process_blog_data(data, output_file):
    # Create DataFrame
    df = create_blog_audit_df(data)

    # Save to Excel
    df.to_excel(output_file, index=False)

    # Apply styling
    style_excel_file(output_file)

    return df

def load_yoast_keywords():
    """Load Yoast keywords data and create URL to keyword mapping"""
    try:
        # Read the Yoast keywords Excel file
        yoast_df = pd.read_excel('resources/yoast-blog-keywords.xlsx', header=None)

        # Create a dictionary mapping URLs to keywords
        url_to_keyword = {}
        for _, row in yoast_df.iterrows():
            url = row[4]  # URL is in column 5 (index 4)
            keyword = row[2]  # Keyword is in column 3 (index 2)
            if pd.notna(url) and pd.notna(keyword):
                url_to_keyword[url.strip()] = keyword.strip()

        return url_to_keyword
    except Exception as e:
        print(f"Error loading Yoast keywords: {e}")
        return {}

def import_performance_data():
    """
    Imports performance data from resources/performance.xlsx and processes it for integration
    with blog audit data.
    """
    try:
        file_path = os.path.join('../resources', 'performance.xlsx')
        performance_df = pd.read_excel(file_path)

        def extract_slug(path):
            if pd.isna(path):
                return ''
            parsed = urlparse(path)
            path_parts = [p for p in parsed.path.strip('/').split('/') if p]
            return path_parts[-1] if path_parts else ''

        performance_df['slug'] = performance_df['Page path + query string'].apply(extract_slug)
        performance_df = performance_df[performance_df['slug'] != '']
        performance_df.set_index('slug', inplace=True)

        return performance_df
    except Exception as e:
        print(f"Warning: Could not load performance data: {str(e)}")
        return None

from ai_analysis import analyze_content_categorization, cost_tracker, analyze_seo, analyze_tone_voice, analyze_quality_brand_fit


def format_seo_data(basic_info, seo_analysis, target_keyword):
    """
    Formats SEO data from content analysis into required structure.

    Args:
        basic_info (dict): Basic content information
        seo_analysis (dict): SEO analysis data
        target_keyword (str): Target keyword from Yoast
    """
    return {
        'current_target_keyword': target_keyword,  # Use the keyword from url_to_keyword mapping
        'meta_description_present': seo_analysis.get('meta_description', {}).get('present', False),
        'h1_present': seo_analysis.get('headings', {}).get('h1_present', False),
        'h2_count': seo_analysis.get('headings', {}).get('h2_count', 0),
        'h3_count': seo_analysis.get('h3_count', 0)
    }


import re

import re

import re
from typing import List, Tuple, Dict

def count_personal_pronouns(text: str) -> Dict[str, any]:
    """
    Counts personal pronouns while excluding those within quotes.
    Additionally, prints the sentences that contain pronouns outside quotes.
    Handles straight (") and curly (“ and ”) double quotes.
    """
    # Define quote pairs: opening quote -> closing quote
    quote_pairs = {
        '"': '"',
        '“': '”',
        '‘': '’',  # Include if handling single curly quotes
    }

    quotes: List[Tuple[int, int]] = []
    current_pos = 0
    text_length = len(text)

    # Step 1: Detect quoted regions
    while current_pos < text_length:
        # Find the next opening quote
        next_quote = None
        next_pos = text_length

        for open_quote in quote_pairs.keys():
            pos = text.find(open_quote, current_pos)
            if pos != -1 and pos < next_pos:
                next_quote = open_quote
                next_pos = pos

        if next_quote is None:
            break  # No more quotes found

        start = next_pos
        end_quote = quote_pairs[next_quote]
        end = text.find(end_quote, start + 1)

        if end == -1:
            # If no closing quote is found, consider the rest of the text as quoted
            end = text_length - 1

        quotes.append((start, end))
        current_pos = end + 1

    def is_in_quotes(pos: int) -> bool:
        """Check if a position is within any quoted region."""
        return any(start <= pos <= end for start, end in quotes)

    # Step 2: Split text into sentences with their positions
    sentence_endings = re.compile(r'([.!?])')  # Regex to identify sentence boundaries
    sentences: List[Tuple[str, int, int]] = []
    start = 0

    for match in sentence_endings.finditer(text):
        end = match.end()
        sentence = text[start:end].strip()
        if sentence:
            sentences.append((sentence, start, end))
        start = end

    # Add any remaining text as the last sentence
    if start < text_length:
        sentence = text[start:].strip()
        if sentence:
            sentences.append((sentence, start, text_length))

    # Step 3: Find pronouns outside quotes
    pronouns = r'\b(I|me|my|mine|myself)\b'
    matches = []
    sentences_with_pronouns: List[str] = []
    seen_sentences: set = set()

    for match in re.finditer(pronouns, text, re.IGNORECASE):
        if not is_in_quotes(match.start()):
            matches.append(match)
            # Step 4: Find the sentence containing this pronoun
            for sentence, s_start, s_end in sentences:
                if s_start <= match.start() < s_end:
                    if sentence not in seen_sentences:
                        sentences_with_pronouns.append(sentence)
                        seen_sentences.add(sentence)
                    break

    return {
        'count': len(matches),
        'found_pronouns': [m.group() for m in matches],
        'quoted_regions': quotes,
        'sentences_with_pronouns': sentences_with_pronouns,
        'flag': len(matches) > 0
    }


def create_blog_audit_df(articles_data):
    """
    Creates a DataFrame containing audit information for multiple articles.
    Includes performance metrics from Google Analytics data.
    """
    # Set up logging
    total_articles = len(articles_data)
    # print(f"\nProcessing {tomake_analysis.pytal_articles} articles...")

    url_to_keyword = load_yoast_keywords()
    all_articles = []
    performance_df = import_performance_data()

    for i, article in enumerate(articles_data, 1):
        cost_tracker.reset()  # Reset cost tracker for new article

        # Safely extract 'content'
        content = article.get('content', '')
        clean_text = clean_content(content)
        print(f"\nArticle {i}/{total_articles}")

        # Safely extract 'basic_info'
        basic_info = article.get('basic_info', {})
        url = basic_info.get('url', 'No URL')
        title = basic_info.get('title', 'No Title')
        publication_date_str = basic_info.get('publication_date')
        modified_date_str = basic_info.get('modified_date')

        # Get pronoun analysis
        pronoun_analysis = count_personal_pronouns(content)

        # Extract target keyword
        target_keyword = url_to_keyword.get(url, 'Not Found')

        # Extract slug for performance data matching
        slug = urlparse(url).path.strip('/').split('/')[-1] if url != 'No URL' else 'No Slug'

        # Safely extract 'multimedia_assessment'
        multimedia = article.get('multimedia_assessment', {})
        header_image = multimedia.get('header_image') or {}

        # Get outdated widget count from multimedia assessment
        outdated_widgets_count = multimedia.get('outdated_widget_count', 0)


        # Safeguard against missing 'width' and 'height'
        try:
            header_width = int(header_image.get('width', 0))
        except (TypeError, ValueError):
            header_width = 0

        try:
            header_height = int(header_image.get('height', 0))
        except (TypeError, ValueError):
            header_height = 0

        header_src = header_image.get('src', 'No Source')
        header_alt = header_image.get('alt', 'No Alt Text')

        # Safely extract 'content_images'
        content_images = multimedia.get('content_images', [])
        content_widths = []
        for img in content_images:
            try:
                width = int(img.get('width', 0))
                content_widths.append(width)
            except (TypeError, ValueError):
                content_widths.append(0)

        min_content_width = min(content_widths) if content_widths else 0

        # Get performance metrics if available
        performance_metrics = {
            'Total Views': 0,
            'Total Users': 0,
            'Total Sessions': 0,
            'Engagement Rate': 0.0,
            'Average Time on Page': 0.0,
            'Bounce Rate': 0.0
        }

        if performance_df is not None and slug in performance_df.index:
            metrics = performance_df.loc[slug]
            performance_metrics = {
                'Total Views': metrics.get('Views', 0),
                'Total Users': metrics.get('Total users', 0),
                'Total Sessions': metrics.get('Sessions', 0),
                'Engagement Rate': metrics.get('Engagement rate', 0.0),
                'Average Time on Page': metrics.get('Average session duration', 0.0),
                'Bounce Rate': metrics.get('Bounce rate', 0.0)
            }

        # Safely extract 'seo_analysis'
        seo_analysis = article.get('seo_analysis', {})
        meta_description = seo_analysis.get('meta_description', {})
        meta_present = meta_description.get('present', False)

        headings = seo_analysis.get('headings', {})
        h1_present = headings.get('h1_present', False)
        h2_count = headings.get('h2_count', 0)
        h3_count = headings.get('h3_count', 0)

        # Parse dates safely
        def parse_date(date_str):
            if date_str:
                try:
                    return datetime.strptime(date_str.split('T')[0], '%Y-%m-%d').strftime('%Y-%m-%d')
                except ValueError:
                    return 'No Date'
            return 'No Date'

        publication_date = parse_date(publication_date_str)
        modified_date = parse_date(modified_date_str)

        print("  ↳ Content categorization...")
        categorization = analyze_content_categorization(clean_text)

        # Format and analyze SEO
        print("  ↳ SEO analysis...")
        seo_data = format_seo_data(basic_info, seo_analysis, target_keyword)
        seo_results = analyze_seo(clean_text, seo_data)

        # Get tone and voice analysis
        print("  ↳ Tone and voice analysis...")
        tone_results = analyze_tone_voice(clean_text)

        # Get quality and brand fit analysis
        print("  ↳ Quality and brand analysis...")
        quality_results = analyze_quality_brand_fit(clean_text)

        # Extract SEO results with error handling
        try:
            keyword_density = float(seo_results.get("Keyword Density", 0))
            keyword_integration = int(seo_results.get("Keyword Integration Score", 0))
            meta_quality = int(seo_results.get("Meta Description Quality Score", 0))
            recommended_keywords = seo_results.get("Recommended New Keywords", [])
            seo_notes = seo_results.get("SEO Notes/Recommendations", "No recommendations available")
        except (ValueError, TypeError, AttributeError) as e:
            print(f"Error processing SEO results for {url}: {str(e)}")
            keyword_density = 0
            keyword_integration = 0
            meta_quality = 0
            recommended_keywords = []
            seo_notes = "Error processing SEO analysis"

        # Extract tone/voice results with error handling
        try:
            challenger_pct = int(tone_results.get("Challenger Percentage", 0))
            supportive_pct = int(tone_results.get("Supportive Percentage", 0))
            natural_score = int(tone_results.get("Natural/Conversational Score", 0))
            authentic_score = int(tone_results.get("Authentic/Approachable Score", 0))
            inclusive_score = int(tone_results.get("Gender-Neutral/Inclusive Score", 0))
            tone_notes = tone_results.get("Tone Notes/Recommendations", "No recommendations available")
        except (ValueError, TypeError, AttributeError) as e:
            print(f"Error processing tone results for {url}: {str(e)}")
            challenger_pct = 0
            supportive_pct = 0
            natural_score = 0
            authentic_score = 0
            inclusive_score = 0
            tone_notes = "Error processing tone analysis"

        # Extract quality/brand results with error handling
        try:
            quality_results = analyze_quality_brand_fit(clean_text)
            if quality_results:  # Check if we got a response
                quality_score = int(quality_results.get("Overall Quality Score", 0))
                brand_alignment = quality_results.get("Brand Alignment", "Needs Work")
                quality_notes = quality_results.get("Quality Notes", "No recommendations available")
                brand_alignment_notes = quality_results.get("Brand Alignment Notes",
                                                            "No brand alignment notes available")
            else:
                raise ValueError("No quality results returned")
        except (ValueError, TypeError, AttributeError) as e:
            print(f"Error processing quality results for {url}: {str(e)}")
            quality_score = 0
            brand_alignment = "Error"
            quality_notes = "Error processing quality analysis"
            brand_alignment_notes = "Error processing brand alignment analysis"

        article_data = {
            # Basic Information
            'Title': title,
            'URL': url,
            'Slug': slug,
            'Publication Date': publication_date,
            'Modified Date': modified_date,
            'Word Count': calculate_word_count(clean_text),
            'Reading Level (Gunning Fog)': round(textstat.gunning_fog(clean_text), 1) if clean_text else 0.0,

            # Quality & Brand Fit
            'Overall Quality Score': quality_score,
            'Topic Relevance': quality_results.get("Topic Relevance", "Error"),  # Add this line
            'Brand Alignment': brand_alignment,
            'Quality Notes/Recommendations': quality_notes,
            'Brand Alignment Notes': brand_alignment_notes,

            # Tone & Voice
            'Challenger Percentage': f"{challenger_pct}%",
            'Supportive Percentage': f"{supportive_pct}%",
            'Natural/Conversational Score': natural_score,
            'Authentic/Approachable Score': authentic_score,
            'Gender-Neutral/Inclusive Score': inclusive_score,
            'Tone Notes/Recommendations': tone_notes,
            'Personal Pronoun Count': pronoun_analysis['count'],

            # SEO Analysis
            'Current Target Keyword': target_keyword,
            'Keyword Density': f"{keyword_density:.2f}%",
            'Keyword Integration Score': keyword_integration,
            'Meta Description Present': 'Yes' if meta_present else 'No',
            'Meta Description Quality Score': meta_quality,
            'H1 Tag Present': 'Yes' if h1_present else 'No',
            'H2 Tags': h2_count,
            'H3 Tags': h3_count,
            'Recommended New Keywords': ' | '.join(recommended_keywords) if recommended_keywords else 'None',
            'SEO Notes/Recommendations': seo_notes,

            # Update Multimedia Assessment section
            'Number of Images': multimedia.get('total_image_count', 0),
            'Header Image Width': header_width,
            'Header Image Height': header_height,
            'Header Image Src': header_src,
            'Header Image Alt': header_alt,
            'Minimum Content Image Width': min_content_width,
            'Outdated Widgets Count': outdated_widgets_count,  # Add this field

            # 'Unmodified Stock Images Flag': 'TBD',

            # Update Content Categorization section with AI results
            'Primary Category': categorization["Primary Category"],
            'Solution Topic': categorization["Solution Topic"],
            'Use Case': categorization["Use Case"],
            'Customer Journey Stage': categorization["Customer Journey Stage"],
            'CMO Priority': categorization["CMO Priority"],
            'Marketing Activity Type': categorization["Marketing Activity Type"],
            'Target Audience': categorization["Target Audience"],

            # Performance Metrics (from Google Analytics)
            **performance_metrics,

            'API Cost': f"${cost_tracker.cost:.4f}",  # Add cost as last field
        }

        print(f"  ✓ Complete (Cost: ${cost_tracker.cost:.4f})")

        all_articles.append(article_data)

    df = pd.DataFrame(all_articles)

    # Convert numeric columns to appropriate types
    numeric_columns = ['Total Views', 'Total Users', 'Total Sessions',
                      'Engagement Rate', 'Average Time on Page', 'Bounce Rate']
    for col in numeric_columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Format time-based metrics
    df['Average Time on Page'] = df['Average Time on Page'].round(2)
    df['Engagement Rate'] = df['Engagement Rate'].round(4)
    df['Bounce Rate'] = df['Bounce Rate'].round(4)

    return df



def process_batch(start_index, batch_size, input_file, output_file, last_successful_index=None):
    """Process a batch of articles and append them to existing Excel file or create new one."""

    try:
        # Read the JSON file
        with open(input_file, 'r') as file:
            data = json.load(file)

        # Get all articles
        articles_data = data.get('analyses', {}).get('blog', [])
        if not articles_data:
            print("No articles found in JSON file")
            return last_successful_index or start_index - 1

        # Get the batch of articles to process
        end_index = min(start_index + batch_size, len(articles_data))
        batch_articles = articles_data[start_index:end_index]

        # Load existing resources
        url_to_keyword = load_yoast_keywords()
        performance_df = import_performance_data()

        # Process articles
        all_articles = []
        current_index = start_index

        for article in batch_articles:
            try:
                cost_tracker.reset()

                # Extract content and clean it
                content = article.get('content', '')
                clean_text = clean_content(content)
                print(f"\nProcessing article {current_index}/{len(articles_data)}")

                # Process article (your existing article processing logic)
                basic_info = article.get('basic_info', {})
                url = basic_info.get('url', 'No URL')
                title = basic_info.get('title', 'No Title')

                # Get all the analysis results using your existing functions
                pronoun_analysis = count_personal_pronouns(content)
                categorization = analyze_content_categorization(clean_text)
                seo_data = format_seo_data(basic_info, article.get('seo_analysis', {}),
                                           url_to_keyword.get(url, 'Not Found'))
                seo_results = analyze_seo(clean_text, seo_data)
                tone_results = analyze_tone_voice(clean_text)
                quality_results = analyze_quality_brand_fit(clean_text)

                # Format article data using your existing logic
                article_data = create_article_data_dict(
                    article, clean_text, basic_info, pronoun_analysis,
                    categorization, seo_results, tone_results, quality_results,
                    performance_df, url_to_keyword
                )

                all_articles.append(article_data)
                last_successful_index = current_index
                print(f"  ✓ Complete (Cost: ${cost_tracker.cost:.4f})")

            except Exception as e:
                print(f"Error processing article at index {current_index}: {str(e)}")
                return last_successful_index or (current_index - 1)

            current_index += 1

        # Convert to DataFrame
        df = pd.DataFrame(all_articles)

        # Check if output file exists
        output_path = Path(output_file)
        if not output_path.exists() or start_index == 0:
            # Create new file with headers
            df.to_excel(output_file, index=False)
            style_excel_file(output_file)
        else:
            # Append to existing file
            append_to_excel(df, output_file)

        return last_successful_index or (current_index - 1)

    except Exception as e:
        print(f"Batch processing error: {str(e)}")
        return last_successful_index or (start_index - 1)



def process_content_data(json_data, output_file):
    """
    Process multiple content pieces from the JSON structure and create an Excel report.

    Parameters:
        json_data (dict): Parsed JSON data containing multiple content analyses
        output_file (str): Path to save the Excel output
    """
    all_content = []
    import json
    from textstat import textstat
    import re
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    import pandas as pd
    from datetime import datetime
    from urllib.parse import urlparse
    import textstat
    import os
    from typing import Dict, List, Tuple

    def clean_content(json_input):
        """
        Cleans and converts JSON content to Markdown.

        Parameters:
            json_input (str): JSON string containing the "content" field.

        Returns:
            str: Cleaned Markdown content.
        """
        try:
            # Parse the JSON input
            data = json.loads(json_input)
            content = data.get("content", "")
        except json.JSONDecodeError:
            # If input is not a valid JSON, assume it's raw content
            content = json_input

        # Remove image markers like [CONTENT IMAGE: ...]
        content = re.sub(r'\[CONTENT IMAGE:.*?\]', '', content)

        # Remove source lines like Source: https://...
        content = re.sub(r'Source:\s*https?://\S+', '', content)

        # Convert 'H2:' headers to Markdown '## '
        content = re.sub(r'H2:\s*', '## ', content)

        # Normalize line breaks and remove excessive whitespace
        content = re.sub(r'\n{3,}', '\n\n', content)  # Replace 3+ newlines with 2
        content = re.sub(r'\n\s*\n', '\n\n', content)  # Ensure double newlines for paragraphs
        content = re.sub(r'[ \t]+', ' ', content)  # Replace multiple spaces/tabs with single space

        # Remove any remaining unwanted brackets or markdown artifacts
        content = re.sub(r'\[.*?\]', '', content)

        # Trim leading and trailing whitespace
        content = content.strip()

        return content

    def calculate_word_count(content):
        clean_text = clean_content(content)
        return len(clean_text.split())

    def add_no_match_highlighting(ws):
        """
        Adds conditional highlighting to mark cells containing 'No Clear Match', 'NONE',
        or similar values in red.

        Parameters:
            ws (openpyxl.worksheet.worksheet.Worksheet): The active worksheet
        """
        # Get header row values
        header_row = [cell.value for cell in ws[2]]  # Headers are in row 2

        # Columns to check for "No Clear" values
        target_columns = [
            'Primary Category',
            'Solution Topic',
            'Use Case',
            'Customer Journey Stage',
            'CMO Priority',
            'Marketing Activity Type',
            'Target Audience'
        ]

        # Get column indices for target columns
        column_indices = {col: header_row.index(col) + 1 for col in target_columns if col in header_row}

        # Values to highlight in red
        highlight_values = [
            'No Clear Match',
            'NONE',
            'No Clear Topic',
            'No Clear Activity Type',
            'No Clear Audience'
        ]

        # Red fill for matching cells
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True)

        # Apply highlighting to matching cells
        for row in range(3, ws.max_row + 1):  # Start from data rows
            for col_name, col_idx in column_indices.items():
                cell = ws.cell(row=row, column=col_idx)
                if cell.value in highlight_values:
                    cell.fill = red_fill
                    cell.font = white_font

    def add_conditional_formatting(ws):
        """Add conditional formatting rules to specific columns."""
        # Get header row values and their indices
        header_row = [cell.value for cell in ws[2]]

        # Mapping of columns to their indices
        columns = {
            'Overall Quality Score': header_row.index('Overall Quality Score') + 1,
            'Topic Relevance': header_row.index('Topic Relevance') + 1,
            'Brand Alignment': header_row.index('Brand Alignment') + 1,
            'Natural/Conversational Score': header_row.index('Natural/Conversational Score') + 1,
            'Authentic/Approachable Score': header_row.index('Authentic/Approachable Score') + 1,
            'Gender-Neutral/Inclusive Score': header_row.index('Gender-Neutral/Inclusive Score') + 1,
            'Personal Pronoun Count': header_row.index('Personal Pronoun Count') + 1,
            'Keyword Integration Score': header_row.index('Keyword Integration Score') + 1,
            'Meta Description Quality Score': header_row.index('Meta Description Quality Score') + 1,
            'Outdated Widgets Count': header_row.index('Outdated Widgets Count') + 1,
            'Word Count': header_row.index('Word Count') + 1

        }

        # Define fill patterns
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        yellow_green_fill = PatternFill(start_color='9ACD32', end_color='9ACD32', fill_type='solid')
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

        # Add ColorScale rules for score columns
        score_columns = ['Overall Quality Score', 'Natural/Conversational Score',
                         'Authentic/Approachable Score', 'Gender-Neutral/Inclusive Score',
                         'Keyword Integration Score', 'Meta Description Quality Score']

        for col_name in score_columns:
            col_letter = get_column_letter(columns[col_name])
            ws.conditional_formatting.add(
                f'{col_letter}3:{col_letter}{ws.max_row}',
                ColorScaleRule(
                    start_type='num', start_value=0, start_color='FF0000',
                    mid_type='num', mid_value=50, mid_color='FFFF00',
                    end_type='num', end_value=100, end_color='00FF00'
                )
            )

        # Add Word Count formatting rules
        word_count_col = get_column_letter(columns['Word Count'])
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=columns['Word Count'])
            try:
                count = int(cell.value)
                if 1000 <= count <= 1200:
                    cell.fill = green_fill
                elif (800 <= count < 1000) or (1201 <= count <= 1400):
                    cell.fill = yellow_fill
                else:
                    cell.fill = orange_fill
            except (ValueError, TypeError):
                pass

        # Add rules for Topic Relevance
        topic_col = get_column_letter(columns['Topic Relevance'])
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=columns['Topic Relevance'])
            if cell.value == "Tangentially Related":
                cell.fill = yellow_fill
            elif cell.value == "Off Topic":
                cell.fill = red_fill

        # Add rules for Brand Alignment
        brand_col = get_column_letter(columns['Brand Alignment'])
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=columns['Brand Alignment'])
            if cell.value == "On Brand":
                cell.fill = green_fill
            elif cell.value == "Mostly on Brand":
                cell.fill = yellow_green_fill
            elif cell.value == "Needs Work":
                cell.fill = yellow_fill
            elif cell.value == "Not on Brand":
                cell.fill = red_fill

        # Add rules for Outdated Widgets Count
        widgets_col = get_column_letter(columns['Outdated Widgets Count'])
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=columns['Outdated Widgets Count'])
            try:
                if int(cell.value) > 0:
                    cell.fill = red_fill
            except (ValueError, TypeError):
                pass

        # Add rules for Personal Pronoun Count
        pronoun_col = get_column_letter(columns['Personal Pronoun Count'])
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=columns['Personal Pronoun Count'])
            try:
                if int(cell.value) > 0:
                    cell.fill = red_fill
            except (ValueError, TypeError):
                pass

    def style_excel_file(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        # Extract title from first row and remove it
        title_cell = ws['A2'].value  # Save title value

        # Define brand colors with lighter shades for subheaders
        section_colors = {
            'Title': {
                'header': '#193661',  # Dark blue
                'subheader': '#C1C9D4'  # Lighter blue
            },
            'Basic Information': {
                'header': '#00babe',  # Teal
                'subheader': '#E5F9F9'  # Lighter teal
            },
            'Quality & Brand Fit': {
                'header': '#e34e64',  # Salmon
                'subheader': '#FFEFEF'  # Lighter salmon
            },
            'Tone & Voice': {
                'header': '#193661',  # Dark blue
                'subheader': '#C1C9D4'  # Lighter blue
            },
            'SEO Analysis': {
                'header': '#00babe',  # Teal
                'subheader': '#E5F9F9'  # Lighter teal
            },
            'Multimedia Assessment': {
                'header': '#e34e64',  # Salmon
                'subheader': '#FFEFEF'  # Lighter salmon
            },
            'Content Categorization': {
                'header': '#193661',  # Dark blue
                'subheader': '#C1C9D4'  # Lighter blue
            },
            'Performance Metrics': {
                'header': '#00babe',  # Teal
                'subheader': '#E5F9F9'  # Lighter teal
            },
            'Cost Analysis': {  # Added missing section
                'header': '#e34e64',  # Salmon
                'subheader': '#FFEFEF'  # Lighter salmon
            }
        }

        # Define sections and their ranges
        sections = {
            'Title': (1, 1),
            'Basic Information': (2, 7),  # Updated to include Modified Date
            'Quality & Brand Fit': (8, 12),
            'Tone & Voice': (13, 19),
            'SEO Analysis': (20, 29),  # Updated for split H2/H3
            'Multimedia Assessment': (30, 36),
            'Content Categorization': (37, 43),
            'Performance Metrics': (44, 49),
            'Cost Analysis': (50, 50)  # New section
        }

        # Freeze top two rows and first column
        ws.freeze_panes = 'B3'

        # Set specific column widths
        ws.column_dimensions['B'].width = 15  # URL column
        for col in range(1, ws.max_column + 1):
            if get_column_letter(col) != 'B':  # All non-URL columns
                ws.column_dimensions[get_column_letter(col)].width = 30

        # # Replace the original column width setting code with:
        # TINY = 4  # Fit 4 characters
        # SMALL = 8  # Fit 8 characters
        # HALF = 15  # Half width
        # DEFAULT = 30  # No change
        #
        # width_map = {
        #     0: DEFAULT, 1: DEFAULT, 2: DEFAULT, 3: HALF, 4: HALF, 5: HALF,
        #     6: TINY, 7: TINY, 8: HALF, 9: DEFAULT, 10: DEFAULT, 11: DEFAULT,
        #     12: TINY, 13: TINY, 14: TINY, 15: TINY, 16: TINY, 17: DEFAULT,
        #     18: TINY, 19: DEFAULT, 20: SMALL, 21: TINY, 22: TINY, 23: TINY,
        #     24: TINY, 25: TINY, 26: TINY, 27: DEFAULT, 28: DEFAULT, 29: TINY,
        #     30: SMALL, 31: SMALL, 32: TINY, 33: DEFAULT, 34: DEFAULT, 35: SMALL,
        #     36: TINY, 37: DEFAULT, 38: DEFAULT, 39: DEFAULT, 40: DEFAULT,
        #     41: DEFAULT, 42: DEFAULT, 43: DEFAULT, 44: DEFAULT, 45: TINY,
        #     46: TINY, 47: TINY, 48: TINY, 49: TINY, 50: TINY
        # }
        #
        # for col_idx, width in width_map.items():
        #     column_letter = get_column_letter(col_idx + 1)
        #     ws.column_dimensions[column_letter].width = width

        # Style metric headers (second row)
        row2 = ws[1]  # Get the second row (index 1)
        for col_idx, cell in enumerate(row2, 1):
            # Find which section this column belongs to
            for section, (start, end) in sections.items():
                if start <= col_idx <= end:
                    subheader_color = section_colors[section]['subheader'].replace('#', '')
                    cell.fill = PatternFill(start_color=subheader_color,
                                            end_color=subheader_color,
                                            fill_type='solid')
                    cell.font = Font(bold=True, color='444444')
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center',
                                                               vertical='center',
                                                               wrap_text=True)
                    break

        # Add and style section headers
        ws.insert_rows(1)
        current_col = 1
        for section, (start, end) in sections.items():
            if current_col <= end:
                cell = ws.cell(row=1, column=current_col)
                cell.value = section
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=section_colors[section]['header'].replace('#', ''),
                                        end_color=section_colors[section]['header'].replace('#', ''),
                                        fill_type='solid')
                cell.alignment = openpyxl.styles.Alignment(horizontal='center',
                                                           vertical='center',
                                                           wrap_text=True)

                if start != end:
                    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=end)
                current_col = end + 1

        # Get column indices for conditional formatting
        header_row = [cell.value for cell in ws[2]]  # Get header row values
        reading_level_col = header_row.index('Reading Level (Gunning Fog)') + 1
        header_image_width_col = header_row.index('Header Image Width') + 1

        # Add conditional formatting for Word Count and Reading Level

        ws.conditional_formatting.add(
            f'{get_column_letter(reading_level_col)}3:{get_column_letter(reading_level_col)}{ws.max_row}',
            ColorScaleRule(
                start_type='min',
                start_color='FF0000',  # Red
                mid_type='percentile',
                mid_value=50,
                mid_color='FFFF00',  # Yellow
                end_type='max',
                end_color='00FF00'  # Green
            )
        )

        # Add conditional formatting for Header Image Width
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=header_image_width_col)
            try:
                width = int(cell.value)
                if width >= 1200:
                    cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
                elif width >= 800:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
                else:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
            except (ValueError, TypeError):
                pass

        # Format data cells (including URL special formatting)
        for row in ws.iter_rows(min_row=3):  # Start from data rows
            for cell in row:
                # Special formatting for URL column
                if cell.column == 2:  # URL column (B)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='left',
                                                               vertical='center',
                                                               wrap_text=False)
                    cell.font = Font(color='0563C1', underline='single')
                    if cell.value and isinstance(cell.value, str):
                        cell.hyperlink = cell.value
                else:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='left',
                                                               vertical='center',
                                                               wrap_text=True)
                    cell.font = Font(color='444444')

        # Add alternating row colors for better readability
        for row in range(3, ws.max_row + 1):
            fill_color = 'F7F9FB' if row % 2 == 0 else 'FFFFFF'
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if not cell.fill.start_color.rgb:  # Only apply if no other fill color is present
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        # Add thin borders to all cells
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='E3E3E3'),
            right=openpyxl.styles.Side(style='thin', color='E3E3E3'),
            top=openpyxl.styles.Side(style='thin', color='E3E3E3'),
            bottom=openpyxl.styles.Side(style='thin', color='E3E3E3')
        )

        # Add the new highlighting for "No Clear Match" values
        add_no_match_highlighting(ws)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        cost_col = header_row.index('API Cost') + 1
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=cost_col)
            cell.number_format = '$#,##0.00000'

        add_conditional_formatting(ws)

        # Restore title in its new position
        ws['A3'] = title_cell

        wb.save(filename)

    def process_blog_data(data, output_file):
        # Create DataFrame
        df = create_blog_audit_df(data)

        # Save to Excel
        df.to_excel(output_file, index=False)

        # Apply styling
        style_excel_file(output_file)

        return df

    def load_yoast_keywords():
        """Load Yoast keywords data and create URL to keyword mapping"""
        try:
            # Read the Yoast keywords Excel file
            yoast_df = pd.read_excel('resources/yoast-blog-keywords.xlsx', header=None)

            # Create a dictionary mapping URLs to keywords
            url_to_keyword = {}
            for _, row in yoast_df.iterrows():
                url = row[4]  # URL is in column 5 (index 4)
                keyword = row[2]  # Keyword is in column 3 (index 2)
                if pd.notna(url) and pd.notna(keyword):
                    url_to_keyword[url.strip()] = keyword.strip()

            return url_to_keyword
        except Exception as e:
            print(f"Error loading Yoast keywords: {e}")
            return {}

    def import_performance_data():
        """
        Imports performance data from resources/performance.xlsx and processes it for integration
        with blog audit data.
        """
        try:
            file_path = os.path.join('../resources', 'performance.xlsx')
            performance_df = pd.read_excel(file_path)

            def extract_slug(path):
                if pd.isna(path):
                    return ''
                parsed = urlparse(path)
                path_parts = [p for p in parsed.path.strip('/').split('/') if p]
                return path_parts[-1] if path_parts else ''

            performance_df['slug'] = performance_df['Page path + query string'].apply(extract_slug)
            performance_df = performance_df[performance_df['slug'] != '']
            performance_df.set_index('slug', inplace=True)

            return performance_df
        except Exception as e:
            print(f"Warning: Could not load performance data: {str(e)}")
            return None

    from ai_analysis import analyze_content_categorization, cost_tracker, analyze_seo, analyze_tone_voice, \
        analyze_quality_brand_fit

    def format_seo_data(basic_info, seo_analysis, target_keyword):
        """
        Formats SEO data from content analysis into required structure.

        Args:
            basic_info (dict): Basic content information
            seo_analysis (dict): SEO analysis data
            target_keyword (str): Target keyword from Yoast
        """
        return {
            'current_target_keyword': target_keyword,  # Use the keyword from url_to_keyword mapping
            'meta_description_present': seo_analysis.get('meta_description', {}).get('present', False),
            'h1_present': seo_analysis.get('headings', {}).get('h1_present', False),
            'h2_count': seo_analysis.get('headings', {}).get('h2_count', 0),
            'h3_count': seo_analysis.get('h3_count', 0)
        }

    import re

    import re

    import re
    from typing import List, Tuple, Dict

    def count_personal_pronouns(text: str) -> Dict[str, any]:
        """
        Counts personal pronouns while excluding those within quotes.
        Additionally, prints the sentences that contain pronouns outside quotes.
        Handles straight (") and curly (“ and ”) double quotes.
        """
        # Define quote pairs: opening quote -> closing quote
        quote_pairs = {
            '"': '"',
            '“': '”',
            '‘': '’',  # Include if handling single curly quotes
        }

        quotes: List[Tuple[int, int]] = []
        current_pos = 0
        text_length = len(text)

        # Step 1: Detect quoted regions
        while current_pos < text_length:
            # Find the next opening quote
            next_quote = None
            next_pos = text_length

            for open_quote in quote_pairs.keys():
                pos = text.find(open_quote, current_pos)
                if pos != -1 and pos < next_pos:
                    next_quote = open_quote
                    next_pos = pos

            if next_quote is None:
                break  # No more quotes found

            start = next_pos
            end_quote = quote_pairs[next_quote]
            end = text.find(end_quote, start + 1)

            if end == -1:
                # If no closing quote is found, consider the rest of the text as quoted
                end = text_length - 1

            quotes.append((start, end))
            current_pos = end + 1

        def is_in_quotes(pos: int) -> bool:
            """Check if a position is within any quoted region."""
            return any(start <= pos <= end for start, end in quotes)

        # Step 2: Split text into sentences with their positions
        sentence_endings = re.compile(r'([.!?])')  # Regex to identify sentence boundaries
        sentences: List[Tuple[str, int, int]] = []
        start = 0

        for match in sentence_endings.finditer(text):
            end = match.end()
            sentence = text[start:end].strip()
            if sentence:
                sentences.append((sentence, start, end))
            start = end

        # Add any remaining text as the last sentence
        if start < text_length:
            sentence = text[start:].strip()
            if sentence:
                sentences.append((sentence, start, text_length))

        # Step 3: Find pronouns outside quotes
        pronouns = r'\b(I|me|my|mine|myself)\b'
        matches = []
        sentences_with_pronouns: List[str] = []
        seen_sentences: set = set()

        for match in re.finditer(pronouns, text, re.IGNORECASE):
            if not is_in_quotes(match.start()):
                matches.append(match)
                # Step 4: Find the sentence containing this pronoun
                for sentence, s_start, s_end in sentences:
                    if s_start <= match.start() < s_end:
                        if sentence not in seen_sentences:
                            sentences_with_pronouns.append(sentence)
                            seen_sentences.add(sentence)
                        break

        return {
            'count': len(matches),
            'found_pronouns': [m.group() for m in matches],
            'quoted_regions': quotes,
            'sentences_with_pronouns': sentences_with_pronouns,
            'flag': len(matches) > 0
        }

    def create_blog_audit_df(articles_data):
        """
        Creates a DataFrame containing audit information for multiple articles.
        Includes performance metrics from Google Analytics data.
        """
        # Set up logging
        total_articles = len(articles_data)
        # print(f"\nProcessing {tomake_analysis.pytal_articles} articles...")

        url_to_keyword = load_yoast_keywords()
        all_articles = []
        performance_df = import_performance_data()

        for i, article in enumerate(articles_data, 1):
            cost_tracker.reset()  # Reset cost tracker for new article

            # Safely extract 'content'
            content = article.get('content', '')
            clean_text = clean_content(content)
            print(f"\nArticle {i}/{total_articles}")

            # Safely extract 'basic_info'
            basic_info = article.get('basic_info', {})
            url = basic_info.get('url', 'No URL')
            title = basic_info.get('title', 'No Title')
            publication_date_str = basic_info.get('publication_date')
            modified_date_str = basic_info.get('modified_date')

            # Get pronoun analysis
            pronoun_analysis = count_personal_pronouns(content)

            # Extract target keyword
            target_keyword = url_to_keyword.get(url, 'Not Found')

            # Extract slug for performance data matching
            slug = urlparse(url).path.strip('/').split('/')[-1] if url != 'No URL' else 'No Slug'

            # Safely extract 'multimedia_assessment'
            multimedia = article.get('multimedia_assessment', {})
            header_image = multimedia.get('header_image') or {}

            # Get outdated widget count from multimedia assessment
            outdated_widgets_count = multimedia.get('outdated_widget_count', 0)

            # Safeguard against missing 'width' and 'height'
            try:
                header_width = int(header_image.get('width', 0))
            except (TypeError, ValueError):
                header_width = 0

            try:
                header_height = int(header_image.get('height', 0))
            except (TypeError, ValueError):
                header_height = 0

            header_src = header_image.get('src', 'No Source')
            header_alt = header_image.get('alt', 'No Alt Text')

            # Safely extract 'content_images'
            content_images = multimedia.get('content_images', [])
            content_widths = []
            for img in content_images:
                try:
                    width = int(img.get('width', 0))
                    content_widths.append(width)
                except (TypeError, ValueError):
                    content_widths.append(0)

            min_content_width = min(content_widths) if content_widths else 0

            # Get performance metrics if available
            performance_metrics = {
                'Total Views': 0,
                'Total Users': 0,
                'Total Sessions': 0,
                'Engagement Rate': 0.0,
                'Average Time on Page': 0.0,
                'Bounce Rate': 0.0
            }

            if performance_df is not None and slug in performance_df.index:
                metrics = performance_df.loc[slug]
                performance_metrics = {
                    'Total Views': metrics.get('Views', 0),
                    'Total Users': metrics.get('Total users', 0),
                    'Total Sessions': metrics.get('Sessions', 0),
                    'Engagement Rate': metrics.get('Engagement rate', 0.0),
                    'Average Time on Page': metrics.get('Average session duration', 0.0),
                    'Bounce Rate': metrics.get('Bounce rate', 0.0)
                }

            # Safely extract 'seo_analysis'
            seo_analysis = article.get('seo_analysis', {})
            meta_description = seo_analysis.get('meta_description', {})
            meta_present = meta_description.get('present', False)

            headings = seo_analysis.get('headings', {})
            h1_present = headings.get('h1_present', False)
            h2_count = headings.get('h2_count', 0)
            h3_count = headings.get('h3_count', 0)

            # Parse dates safely
            def parse_date(date_str):
                if date_str:
                    try:
                        return datetime.strptime(date_str.split('T')[0], '%Y-%m-%d').strftime('%Y-%m-%d')
                    except ValueError:
                        return 'No Date'
                return 'No Date'

            publication_date = parse_date(publication_date_str)
            modified_date = parse_date(modified_date_str)

            print("  ↳ Content categorization...")
            categorization = analyze_content_categorization(clean_text)

            # Format and analyze SEO
            print("  ↳ SEO analysis...")
            seo_data = format_seo_data(basic_info, seo_analysis, target_keyword)
            seo_results = analyze_seo(clean_text, seo_data)

            # Get tone and voice analysis
            print("  ↳ Tone and voice analysis...")
            tone_results = analyze_tone_voice(clean_text)

            # Get quality and brand fit analysis
            print("  ↳ Quality and brand analysis...")
            quality_results = analyze_quality_brand_fit(clean_text)

            # Extract SEO results with error handling
            try:
                keyword_density = float(seo_results.get("Keyword Density", 0))
                keyword_integration = int(seo_results.get("Keyword Integration Score", 0))
                meta_quality = int(seo_results.get("Meta Description Quality Score", 0))
                recommended_keywords = seo_results.get("Recommended New Keywords", [])
                seo_notes = seo_results.get("SEO Notes/Recommendations", "No recommendations available")
            except (ValueError, TypeError, AttributeError) as e:
                print(f"Error processing SEO results for {url}: {str(e)}")
                keyword_density = 0
                keyword_integration = 0
                meta_quality = 0
                recommended_keywords = []
                seo_notes = "Error processing SEO analysis"

            # Extract tone/voice results with error handling
            try:
                challenger_pct = int(tone_results.get("Challenger Percentage", 0))
                supportive_pct = int(tone_results.get("Supportive Percentage", 0))
                natural_score = int(tone_results.get("Natural/Conversational Score", 0))
                authentic_score = int(tone_results.get("Authentic/Approachable Score", 0))
                inclusive_score = int(tone_results.get("Gender-Neutral/Inclusive Score", 0))
                tone_notes = tone_results.get("Tone Notes/Recommendations", "No recommendations available")
            except (ValueError, TypeError, AttributeError) as e:
                print(f"Error processing tone results for {url}: {str(e)}")
                challenger_pct = 0
                supportive_pct = 0
                natural_score = 0
                authentic_score = 0
                inclusive_score = 0
                tone_notes = "Error processing tone analysis"

            # Extract quality/brand results with error handling
            try:
                quality_results = analyze_quality_brand_fit(clean_text)
                if quality_results:  # Check if we got a response
                    quality_score = int(quality_results.get("Overall Quality Score", 0))
                    brand_alignment = quality_results.get("Brand Alignment", "Needs Work")
                    quality_notes = quality_results.get("Quality Notes", "No recommendations available")
                    brand_alignment_notes = quality_results.get("Brand Alignment Notes",
                                                                "No brand alignment notes available")
                else:
                    raise ValueError("No quality results returned")
            except (ValueError, TypeError, AttributeError) as e:
                print(f"Error processing quality results for {url}: {str(e)}")
                quality_score = 0
                brand_alignment = "Error"
                quality_notes = "Error processing quality analysis"
                brand_alignment_notes = "Error processing brand alignment analysis"

            article_data = {
                # Basic Information
                'Title': title,
                'URL': url,
                'Slug': slug,
                'Publication Date': publication_date,
                'Modified Date': modified_date,
                'Word Count': calculate_word_count(clean_text),
                'Reading Level (Gunning Fog)': round(textstat.gunning_fog(clean_text), 1) if clean_text else 0.0,

                # Quality & Brand Fit
                'Overall Quality Score': quality_score,
                'Topic Relevance': quality_results.get("Topic Relevance", "Error"),  # Add this line
                'Brand Alignment': brand_alignment,
                'Quality Notes/Recommendations': quality_notes,
                'Brand Alignment Notes': brand_alignment_notes,

                # Tone & Voice
                'Challenger Percentage': f"{challenger_pct}%",
                'Supportive Percentage': f"{supportive_pct}%",
                'Natural/Conversational Score': natural_score,
                'Authentic/Approachable Score': authentic_score,
                'Gender-Neutral/Inclusive Score': inclusive_score,
                'Tone Notes/Recommendations': tone_notes,
                'Personal Pronoun Count': pronoun_analysis['count'],

                # SEO Analysis
                'Current Target Keyword': target_keyword,
                'Keyword Density': f"{keyword_density:.2f}%",
                'Keyword Integration Score': keyword_integration,
                'Meta Description Present': 'Yes' if meta_present else 'No',
                'Meta Description Quality Score': meta_quality,
                'H1 Tag Present': 'Yes' if h1_present else 'No',
                'H2 Tags': h2_count,
                'H3 Tags': h3_count,
                'Recommended New Keywords': ' | '.join(recommended_keywords) if recommended_keywords else 'None',
                'SEO Notes/Recommendations': seo_notes,

                # Update Multimedia Assessment section
                'Number of Images': multimedia.get('total_image_count', 0),
                'Header Image Width': header_width,
                'Header Image Height': header_height,
                'Header Image Src': header_src,
                'Header Image Alt': header_alt,
                'Minimum Content Image Width': min_content_width,
                'Outdated Widgets Count': outdated_widgets_count,  # Add this field

                # 'Unmodified Stock Images Flag': 'TBD',

                # Update Content Categorization section with AI results
                'Primary Category': categorization["Primary Category"],
                'Solution Topic': categorization["Solution Topic"],
                'Use Case': categorization["Use Case"],
                'Customer Journey Stage': categorization["Customer Journey Stage"],
                'CMO Priority': categorization["CMO Priority"],
                'Marketing Activity Type': categorization["Marketing Activity Type"],
                'Target Audience': categorization["Target Audience"],

                # Performance Metrics (from Google Analytics)
                **performance_metrics,

                'API Cost': f"${cost_tracker.cost:.4f}",  # Add cost as last field
            }

            print(f"  ✓ Complete (Cost: ${cost_tracker.cost:.4f})")

            all_articles.append(article_data)

        df = pd.DataFrame(all_articles)

        # Convert numeric columns to appropriate types
        numeric_columns = ['Total Views', 'Total Users', 'Total Sessions',
                           'Engagement Rate', 'Average Time on Page', 'Bounce Rate']
        for col in numeric_columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Format time-based metrics
        df['Average Time on Page'] = df['Average Time on Page'].round(2)
        df['Engagement Rate'] = df['Engagement Rate'].round(4)
        df['Bounce Rate'] = df['Bounce Rate'].round(4)

        return df

    def process_content_data(json_data, output_file):
        """
        Process multiple content pieces from the JSON structure and create an Excel report.

        Parameters:
            json_data (dict): Parsed JSON data containing multiple content analyses
            output_file (str): Path to save the Excel output
        """
        all_content = []

        # Process each content type
        for content_type, content_list in json_data['analyses'].items():
            # Skip empty content types
            if not content_list:
                continue

            # Add each piece of content to the list
            for content in content_list:
                content['content_type'] = content_type  # Add content type to the data
                all_content.append(content)

        # Create DataFrame with all content
        df = create_blog_audit_df(all_content)

        # Save to Excel
        df.to_excel(output_file, index=False)

        # Apply styling
        style_excel_file(output_file)

        return df

    # Example usage
    if __name__ == "__main__":
        input_file = '../output/blog.json'
        output_file = '../output/blog_audit.xlsx'

        # Read the JSON file
        with open(input_file, 'r') as file:
            data = json.load(file)

        # Process the data
        df = process_content_data(data, output_file)
        print(f"Excel file created: {output_file}")

    # Process each content type
    for content_type, content_list in json_data['analyses'].items():
        # Skip empty content types
        if not content_list:
            continue

        # Add each piece of content to the list
        for content in content_list:
            content['content_type'] = content_type  # Add content type to the data
            all_content.append(content)

    # Create DataFrame with all content
    df = create_blog_audit_df(all_content)

    # Save to Excel
    df.to_excel(output_file, index=False)

    # Apply styling
    style_excel_file(output_file)

    return df


def process_multiple_batches(start_index, total_size, batch_size, input_file, output_file):
    """Process multiple batches of articles."""

    last_successful_index = None
    remaining_size = total_size
    current_start = start_index

    while remaining_size > 0:
        current_batch_size = min(batch_size, remaining_size)
        print(f"\nProcessing batch: {current_start} to {current_start + current_batch_size - 1}")

        last_successful_index = process_batch(
            current_start,
            current_batch_size,
            input_file,
            output_file,
            last_successful_index
        )

        if last_successful_index < current_start + current_batch_size - 1:
            print(f"\nProcessing stopped at index {last_successful_index}")
            print("Please restart processing from index", last_successful_index + 1)
            break

        current_start += current_batch_size
        remaining_size -= current_batch_size
        print(f"Completed batch. Last successful index: {last_successful_index}")


def create_article_data_dict(article, clean_text, basic_info, pronoun_analysis, categorization,
                             seo_results, tone_results, quality_results, performance_df, url_to_keyword):
    """Create the standardized article data dictionary."""

    # Extract all the necessary data using your existing logic
    url = basic_info.get('url', 'No URL')
    title = basic_info.get('title', 'No Title')
    slug = urlparse(url).path.strip('/').split('/')[-1] if url != 'No URL' else 'No Slug'

    # Get performance metrics
    performance_metrics = get_performance_metrics(slug, performance_df)

    # Extract all other necessary data using your existing logic for multimedia, dates, etc.
    multimedia = article.get('multimedia_assessment', {})
    header_image = multimedia.get('header_image', {})

    # Create and return the full article data dictionary
    return {
        # Your existing dictionary structure with all the fields
        'Title': title,
        'URL': url,
        'Slug': slug,
        'Publication Date': parse_date(basic_info.get('publication_date')),
        'Modified Date': parse_date(basic_info.get('modified_date')),
        'Word Count': calculate_word_count(clean_text),
        'Reading Level (Gunning Fog)': round(textstat.gunning_fog(clean_text), 1) if clean_text else 0.0,
        # ... (rest of your existing dictionary structure)
        **performance_metrics,
        'API Cost': f"${cost_tracker.cost:.4f}"
    }

# Example usage
if __name__ == "__main__":
    input_file = '../output/blog.json'
    output_file = '../output/blog_audit.xlsx'

    # Read the JSON file
    with open(input_file, 'r') as file:
        data = json.load(file)

    # Process the data
    df = process_content_data(data, output_file)
    print(f"Excel file created: {output_file}")
