import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import glob

class ExcelStyler:
    def __init__(self):
        # Define brand colors with lighter shades for subheaders
        self.section_colors = {
            'Title': {'header': '#193661', 'subheader': '#C1C9D4'},
            'Basic Information': {'header': '#00babe', 'subheader': '#E5F9F9'},
            'Quality & Brand Fit': {'header': '#e34e64', 'subheader': '#FFEFEF'},
            'Tone & Voice': {'header': '#193661', 'subheader': '#C1C9D4'},
            'SEO Analysis': {'header': '#00babe', 'subheader': '#E5F9F9'},
            'Multimedia Assessment': {'header': '#e34e64', 'subheader': '#FFEFEF'},
            'Content Categorization': {'header': '#193661', 'subheader': '#C1C9D4'},
            'Performance Metrics': {'header': '#00babe', 'subheader': '#E5F9F9'},
            'Cost Analysis': {'header': '#e34e64', 'subheader': '#FFEFEF'}
        }
        # Define sections and their column ranges
        self.sections = {
            'Title': (1, 1),
            'Basic Information': (2, 7),
            'Quality & Brand Fit': (8, 12),
            'Tone & Voice': (13, 19),
            'SEO Analysis': (20, 29),
            'Multimedia Assessment': (30, 36),
            'Content Categorization': (37, 43),
            'Performance Metrics': (44, 49),
            'Cost Analysis': (50, 50)
        }
        # Define common styles
        self.thin_border = Border(
            left=Side(style='thin', color='E3E3E3'),
            right=Side(style='thin', color='E3E3E3'),
            top=Side(style='thin', color='E3E3E3'),
            bottom=Side(style='thin', color='E3E3E3')
        )

    def apply_full_styling(self, filename: str):
        """Apply all styles to the Excel file."""
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        self._setup_basic_formatting(ws)
        self._add_section_headers(ws)
        self._apply_base_styles(ws)  # New method for base styles
        self._add_conditional_formatting(ws)  # Move conditional formatting after base styles
        self._add_no_match_highlighting(ws)
        self._apply_column_widths(ws)

        wb.save(filename)

    def _apply_base_styles(self, ws):
        """Apply base styles including alternating row colors."""
        # Style metric headers (second row)
        for col_idx, cell in enumerate(ws[2], 1):
            for section, (start, end) in self.sections.items():
                if start <= col_idx <= end:
                    subheader_color = self.section_colors[section]['subheader'].replace('#', '')
                    cell.fill = PatternFill(start_color=subheader_color,
                                            end_color=subheader_color,
                                            fill_type='solid')
                    cell.font = Font(bold=True, color='444444')
                    cell.alignment = Alignment(horizontal='center',
                                               vertical='center',
                                               wrap_text=True)
                    break

        # Apply base styles to data cells
        for row in range(3, ws.max_row + 1):
            fill_color = 'F7F9FB' if row % 2 == 0 else 'FFFFFF'
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                # Special formatting for URL column
                if col == 2:  # URL column (B)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                    cell.font = Font(color='0563C1', underline='single')
                    if cell.value and isinstance(cell.value, str):
                        cell.hyperlink = cell.value
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.font = Font(color='444444')

                # Only apply alternating row colors to cells that won't have conditional formatting
                if not self._will_have_conditional_formatting(cell, col, header_row=[c.value for c in ws[2]]):
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

                cell.border = self.thin_border

    def _will_have_conditional_formatting(self, cell, col_idx, header_row):
        """Determine if a cell will receive conditional formatting."""
        if not header_row[col_idx - 1]:  # Check if header exists
            return False

        column_name = header_row[col_idx - 1]

        # List of columns that will receive conditional formatting
        conditional_columns = {
            'Overall Quality Score',
            'Natural/Conversational Score',
            'Authentic/Approachable Score',
            'Gender-Neutral/Inclusive Score',
            'Keyword Integration Score',
            'Meta Description Quality Score',
            'Reading Level (Gunning Fog)',
            'Word Count',
            'Topic Relevance',
            'Brand Alignment',
            'Outdated Widgets Count',
            'Personal Pronoun Count',
            'Header Image Width'
        }

        return column_name in conditional_columns

    def _setup_basic_formatting(self, ws):
        """Apply basic worksheet formatting."""
        ws.freeze_panes = 'B3'

    def _add_section_headers(self, ws):
        """Add and style section headers."""
        ws.insert_rows(1)
        current_col = 1
        for section, (start, end) in self.sections.items():
            if current_col <= end:
                cell = ws.cell(row=1, column=current_col)
                cell.value = section
                cell.font = Font(bold=True, color='FFFFFF')
                header_color = self.section_colors[section]['header'].replace('#', '')
                cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if start != end:
                    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=end)
                current_col = end + 1

    def _style_data_cells(self, ws):
        """Style all data cells including alternating row colors."""
        # Style metric headers (second row)
        for col_idx, cell in enumerate(ws[2], 1):
            for section, (start, end) in self.sections.items():
                if start <= col_idx <= end:
                    subheader_color = self.section_colors[section]['subheader'].replace('#', '')
                    cell.fill = PatternFill(start_color=subheader_color,
                                            end_color=subheader_color,
                                            fill_type='solid')
                    cell.font = Font(bold=True, color='444444')
                    cell.alignment = Alignment(horizontal='center',
                                               vertical='center',
                                               wrap_text=True)
                    break

        # Style data cells
        for row in range(3, ws.max_row + 1):
            fill_color = 'F7F9FB' if row % 2 == 0 else 'FFFFFF'
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                # Special formatting for URL column
                if col == 2:  # URL column (B)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                    cell.font = Font(color='0563C1', underline='single')
                    if cell.value and isinstance(cell.value, str):
                        cell.hyperlink = cell.value
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.font = Font(color='444444')
                # Apply fill color
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.border = self.thin_border

    def _add_conditional_formatting(self, ws):
        """Add conditional formatting rules to specific columns."""
        header_row = [cell.value for cell in ws[2]]
        columns = {col_name: header_row.index(col_name) + 1 for col_name in header_row if col_name}

        # Define fill patterns
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
        yellow_green_fill = PatternFill(start_color='FF9ACD32', end_color='FF9ACD32', fill_type='solid')
        orange_fill = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')

        # Add ColorScale rules for score columns
        score_columns = [
            'Overall Quality Score',
            'Natural/Conversational Score',
            'Authentic/Approachable Score',
            'Gender-Neutral/Inclusive Score',
            'Keyword Integration Score',
            'Meta Description Quality Score'
        ]

        for col_name in score_columns:
            if col_name in columns:
                col_letter = get_column_letter(columns[col_name])
                ws.conditional_formatting.add(
                    f'{col_letter}3:{col_letter}{ws.max_row}',
                    ColorScaleRule(
                        start_type='num', start_value=0, start_color='FFFF0000',
                        mid_type='num', mid_value=50, mid_color='FFFFFF00',
                        end_type='num', end_value=100, end_color='FF00FF00'
                    )
                )

        # Add Reading Level gradient
        if 'Reading Level (Gunning Fog)' in columns:
            col_letter = get_column_letter(columns['Reading Level (Gunning Fog)'])
            ws.conditional_formatting.add(
                f'{col_letter}3:{col_letter}{ws.max_row}',
                ColorScaleRule(
                    start_type='min', start_color='FFFF0000',
                    mid_type='percentile', mid_value=50, mid_color='FFFFFF00',
                    end_type='max', end_color='FF00FF00'
                )
            )

        # Add Word Count formatting
        if 'Word Count' in columns:
            word_count_col = columns['Word Count']
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=word_count_col)
                try:
                    count = int(cell.value)
                    if 1000 <= count <= 1200:
                        cell.fill = green_fill
                    elif (800 <= count < 1000) or (1201 <= count <= 1400):
                        cell.fill = yellow_fill
                    else:
                        cell.fill = orange_fill
                except (ValueError, TypeError):
                    pass

        # Add Brand Alignment formatting
        if 'Brand Alignment' in columns:
            brand_col = columns['Brand Alignment']
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=brand_col)
                if cell.value == "On Brand":
                    cell.fill = green_fill
                elif cell.value == "Mostly on Brand":
                    cell.fill = yellow_green_fill
                elif cell.value == "Needs Work":
                    cell.fill = yellow_fill
                elif cell.value == "Not on Brand":
                    cell.fill = red_fill

        # Add Topic Relevance formatting
        if 'Topic Relevance' in columns:
            topic_col = columns['Topic Relevance']
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=topic_col)
                if cell.value == "Tangentially Related":
                    cell.fill = yellow_fill
                elif cell.value == "Off Topic":
                    cell.fill = red_fill

        # Add warning highlighting for counts
        for col_name in ['Outdated Widgets Count', 'Personal Pronoun Count']:
            if col_name in columns:
                col_idx = columns[col_name]
                for row in range(3, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        if int(cell.value) > 0:
                            cell.fill = red_fill
                    except (ValueError, TypeError):
                        pass

        # Add Header Image Width formatting
        if 'Header Image Width' in columns:
            img_width_col = columns['Header Image Width']
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=img_width_col)
                try:
                    width = int(cell.value)
                    if width >= 1200:
                        cell.fill = green_fill
                    elif width >= 800:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill
                except (ValueError, TypeError):
                    pass

        # Format cost column
        if 'API Cost' in columns:
            cost_col = columns['API Cost']
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=cost_col)
                cell.number_format = '$#,##0.00000'

    def _add_no_match_highlighting(self, ws):
        """Add red highlighting for 'No Clear Match' values."""
        header_row = [cell.value for cell in ws[2]]
        columns = {col_name: header_row.index(col_name) + 1 for col_name in header_row if col_name}
        target_columns = [
            'Primary Category',
            'Solution Topic',
            'Use Case',
            'Customer Journey Stage',
            'CMO Priority',
            'Marketing Activity Type',
            'Target Audience'
        ]
        highlight_values = [
            'No Clear Match',
            'NONE',
            'No Clear Topic',
            'No Clear Activity Type',
            'No Clear Audience'
        ]
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        white_font = Font(color='FFFFFFFF', bold=True)

        for col_name in target_columns:
            if col_name in columns:
                col_idx = columns[col_name]
                for row in range(3, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value in highlight_values:
                        cell.fill = red_fill
                        cell.font = white_font

    def _apply_column_widths(self, ws):
        """Set appropriate column widths."""
        ws.column_dimensions['B'].width = 15  # URL column
        for col in range(1, ws.max_column + 1):
            if get_column_letter(col) != 'B':  # All non-URL columns
                ws.column_dimensions[get_column_letter(col)].width = 30

# Step 1: Combine spreadsheets
def combine_spreadsheets():
    folder_path = "output/backup/*.xlsx"
    files = glob.glob(folder_path)
    all_dataframes = []

    for file in files:
        # Read without specifying header=[0, 1] since we want the original single-row headers
        df = pd.read_excel(file)
        all_dataframes.append(df)

    combined_df = pd.concat(all_dataframes, ignore_index=True)

    output_filename = "../combined_spreadsheets_last.xlsx"
    combined_df.to_excel(output_filename, index=False, engine='openpyxl')

    return output_filename

# Step 2: Apply styling
def main():
    combined_file = combine_spreadsheets()
    styler = ExcelStyler()
    styler.apply_full_styling(combined_file)
    print("Spreadsheets combined and styled successfully!")

if __name__ == "__main__":
    main()