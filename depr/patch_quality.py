import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
from typing import List, Tuple
from depr.ai_analysis import analyze_quality_brand_fit



def load_content_from_json(url: str, json_file: str = 'output/all.json') -> Tuple[str, str]:
    """Find and return content for a URL from the JSON file."""
    try:
        with open(json_file, 'r') as f:
            data = json.load(f)

        # Search through blog entries
        for entry in data['analyses'].get('blog', []):
            if entry.get('url') == url:
                return entry.get('content', ''), entry.get('basic_info', {}).get('title', url)

        print(f"URL not found in JSON file: {url}")
        return '', url

    except Exception as e:
        print(f"Error reading JSON file: {str(e)}")
        return '', url


def analyze_urls(urls: List[str]) -> pd.DataFrame:
    """Analyze multiple URLs and return results as a DataFrame."""
    results = []

    for url in urls:
        print(f"\nAnalyzing {url}...")
        content, title = load_content_from_json(url)

        if not content:
            print(f"No content found for {url}, skipping...")
            continue

        try:
            # Run quality analysis with more detailed error output
            print("Running quality analysis...")
            analysis = analyze_quality_brand_fit(content)

            print("Analysis result:")
            print(f"Quality Score: {analysis.get('Overall Quality Score', 'N/A')}")
            print(f"Topic Relevance: {analysis.get('Topic Relevance', 'N/A')}")
            print(f"Brand Alignment: {analysis.get('Brand Alignment', 'N/A')}")

            # Store whatever results we got, even if partial
            result = {
                "Title": title,
                "URL": url,
                "Overall Quality Score": analysis.get("Overall Quality Score", "n/a"),
                "Topic Relevance": analysis.get("Topic Relevance", "n/a"),
                "Brand Alignment": analysis.get("Brand Alignment", "n/a"),
                "Quality Notes": analysis.get("Quality Notes", "No quality notes available"),
                "Brand Alignment Notes": analysis.get("Brand Alignment Notes", "No brand alignment notes available")
            }
            results.append(result)

        except Exception as e:
            print(f"Error analyzing {url}: {str(e)}")
            # Store error state but don't use default values
            results.append({
                "Title": title,
                "URL": url,
                "Overall Quality Score": "error",
                "Topic Relevance": "error",
                "Brand Alignment": "error",
                "Quality Notes": f"Error during analysis: {str(e)}",
                "Brand Alignment Notes": f"Error during analysis: {str(e)}"
            })

    return pd.DataFrame(results)


def style_excel_output(filename: str):
    """Apply styling to the Excel output."""
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Define styles
    header_fill = PatternFill(start_color="193661", end_color="193661", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='E3E3E3'),
        right=Side(style='thin', color='E3E3E3'),
        top=Side(style='thin', color='E3E3E3'),
        bottom=Side(style='thin', color='E3E3E3')
    )

    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions['A'].width = 40  # Title
    ws.column_dimensions['B'].width = 50  # URL
    ws.column_dimensions['C'].width = 15  # Quality Score
    ws.column_dimensions['D'].width = 15  # Topic Relevance
    ws.column_dimensions['E'].width = 15  # Brand Alignment
    ws.column_dimensions['F'].width = 50  # Quality Notes
    ws.column_dimensions['G'].width = 50  # Brand Alignment Notes

    # Style data cells
    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row):
            cell.border = thin_border
            if idx == 1:  # URL column
                cell.font = Font(color="0563C1", underline="single")
                cell.hyperlink = cell.value

            # Add conditional formatting for Quality Score
            if idx == 2:  # Overall Quality Score column
                try:
                    score = float(cell.value)
                    if score >= 80:
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    elif score >= 60:
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                except (ValueError, TypeError):
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            # Add conditional formatting for Brand Alignment
            if idx == 4:  # Brand Alignment column
                if cell.value == "On Brand":
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif cell.value == "Mostly on Brand":
                    cell.fill = PatternFill(start_color="9ACD32", end_color="9ACD32", fill_type="solid")
                elif cell.value == "Needs Work":
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif cell.value == "Not on Brand":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                elif cell.value.lower() == "error":
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    wb.save(filename)


def debug_response(response, url: str):
    """Helper function to print detailed debug info about API response"""
    print(f"\nDebug info for {url}:")
    print("Raw response type:", type(response))
    print("Raw response:", response)
    if isinstance(response, dict):
        for key, value in response.items():
            print(f"{key}: {type(value)} = {value}")


def main(urls: List[str], output_filename: str = "quality_analysis.xlsx"):
    """Main function to analyze URLs and create styled Excel output."""
    print(f"Starting analysis of {len(urls)} URLs...")

    # Analyze URLs
    df = analyze_urls(urls)

    # Print summary before saving
    print("\nAnalysis Summary:")
    print(f"Total URLs processed: {len(df)}")
    print(f"Successful analyses: {len(df[df['Overall Quality Score'] != 'error'])}")
    print(f"Failed analyses: {len(df[df['Overall Quality Score'] == 'error'])}")

    # Save to Excel
    print(f"\nSaving results to {output_filename}...")
    df.to_excel(output_filename, index=False, engine='openpyxl')

    # Apply styling
    style_excel_output(output_filename)

    print("Analysis complete!")


if __name__ == "__main__":
    urls_to_analyze = [
        'https://act-on.com/learn/blog/how-to-create-a-successful-webinar-start-by-asking-these-7-questions/','https://act-on.com/learn/blog/how-to-create-an-editorial-calendar/','https://act-on.com/learn/blog/how-to-create-an-effective-video-marketing-strategy/','https://act-on.com/learn/blog/how-to-create-an-engaging-and-effective-presentation/','https://act-on.com/learn/blog/how-to-create-an-organic-linkedin-b2b-marketing-strategy/','https://act-on.com/learn/blog/how-to-create-better-content-in-2018/','https://act-on.com/learn/blog/how-to-create-content-that-generates-brand-awareness-and-helps-you-reach-more-leads/','https://act-on.com/learn/blog/how-to-create-content-unicorns-and-ignore-the-donkeys/','https://act-on.com/learn/blog/how-to-create-ctas-that-convert/'
    ]
    main(urls_to_analyze)