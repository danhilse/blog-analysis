import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


class ExcelStyler:
    def __init__(self):
        self.section_colors = {
            'Title': {
                'header': '#193661',  # Dark blue
                'subheader': '#C1C9D4'  # Lighter blue
            },
            'Basic Information': {
                'header': '#00babe',  # Teal
                'subheader': '#E5F9F9'  # Lighter teal
            },
            'Quality & Brand Fit': {
                'header': '#e34e64',  # Salmon
                'subheader': '#FFEFEF'  # Lighter salmon
            },
            'Tone & Voice': {
                'header': '#193661',  # Dark blue
                'subheader': '#C1C9D4'  # Lighter blue
            },
            'SEO Analysis': {
                'header': '#00babe',  # Teal
                'subheader': '#E5F9F9'  # Lighter teal
            },
            'Multimedia Assessment': {
                'header': '#e34e64',  # Salmon
                'subheader': '#FFEFEF'  # Lighter salmon
            },
            'Content Categorization': {
                'header': '#193661',  # Dark blue
                'subheader': '#C1C9D4'  # Lighter blue
            },
            'Performance Metrics': {
                'header': '#00babe',  # Teal
                'subheader': '#E5F9F9'  # Lighter teal
            },
            'Cost Analysis': {
                'header': '#e34e64',  # Salmon
                'subheader': '#FFEFEF'  # Lighter salmon
            }
        }

        self.sections = {
            'Title': (1, 1),
            'Basic Information': (2, 7),
            'Quality & Brand Fit': (8, 12),
            'Tone & Voice': (13, 19),
            'SEO Analysis': (20, 29),
            'Multimedia Assessment': (30, 36),
            'Content Categorization': (37, 43),
            'Performance Metrics': (44, 49),
            'Cost Analysis': (50, 50)
        }

    def add_no_match_highlighting(self, ws):
        """Adds conditional highlighting for 'No Clear Match' values."""
        header_row = [cell.value for cell in ws[2]]

        target_columns = [
            'Primary Category',
            'Solution Topic',
            'Use Case',
            'Customer Journey Stage',
            'CMO Priority',
            'Marketing Activity Type',
            'Target Audience'
        ]

        column_indices = {col: header_row.index(col) + 1 for col in target_columns if col in header_row}
        highlight_values = [
            'No Clear Match',
            'NONE',
            'No Clear Topic',
            'No Clear Activity Type',
            'No Clear Audience'
        ]

        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True)

        for row in range(3, ws.max_row + 1):
            for col_name, col_idx in column_indices.items():
                cell = ws.cell(row=row, column=col_idx)
                if cell.value in highlight_values:
                    cell.fill = red_fill
                    cell.font = white_font

    def add_conditional_formatting(self, ws):
        """Adds conditional formatting to specific columns."""
        header_row = [cell.value for cell in ws[2]]

        columns = {
            'Overall Quality Score': header_row.index('Overall Quality Score') + 1,
            'Topic Relevance': header_row.index('Topic Relevance') + 1,
            'Brand Alignment': header_row.index('Brand Alignment') + 1,
            'Natural/Conversational Score': header_row.index('Natural/Conversational Score') + 1,
            'Authentic/Approachable Score': header_row.index('Authentic/Approachable Score') + 1,
            'Gender-Neutral/Inclusive Score': header_row.index('Gender-Neutral/Inclusive Score') + 1,
            'Personal Pronoun Count': header_row.index('Personal Pronoun Count') + 1,
            'Keyword Integration Score': header_row.index('Keyword Integration Score') + 1,
            'Meta Description Quality Score': header_row.index('Meta Description Quality Score') + 1,
            'Outdated Widgets Count': header_row.index('Outdated Widgets Count') + 1,
            'Word Count': header_row.index('Word Count') + 1
        }

        # Add score-based conditional formatting
        score_columns = ['Overall Quality Score', 'Natural/Conversational Score',
                         'Authentic/Approachable Score', 'Gender-Neutral/Inclusive Score',
                         'Keyword Integration Score', 'Meta Description Quality Score']

        for col_name in score_columns:
            col_letter = get_column_letter(columns[col_name])
            ws.conditional_formatting.add(
                f'{col_letter}3:{col_letter}{ws.max_row}',
                ColorScaleRule(
                    start_type='num', start_value=0, start_color='FF0000',
                    mid_type='num', mid_value=50, mid_color='FFFF00',
                    end_type='num', end_value=100, end_color='00FF00'
                )
            )

        # Add word count formatting
        word_count_col = get_column_letter(columns['Word Count'])
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=columns['Word Count'])
            try:
                count = int(cell.value)
                if 1000 <= count <= 1200:
                    cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                elif (800 <= count < 1000) or (1201 <= count <= 1400):
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            except (ValueError, TypeError):
                pass

        # Add other conditional formatting rules here...
        # (Topic Relevance, Brand Alignment, etc.)

    def style_excel_file(self, filename):
        """Main function to apply all styling to the Excel file."""
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        # Extract and save title
        title_cell = ws['A2'].value

        # Freeze panes
        ws.freeze_panes = 'B3'

        # Set column widths
        ws.column_dimensions['B'].width = 15
        for col in range(1, ws.max_column + 1):
            if get_column_letter(col) != 'B':
                ws.column_dimensions[get_column_letter(col)].width = 30

        # Style headers and add sections
        self._style_headers(ws)
        self._add_sections(ws)

        # Add conditional formatting
        self.add_no_match_highlighting(ws)
        self.add_conditional_formatting(ws)

        # Format data cells
        self._format_data_cells(ws)

        # Add alternating row colors
        self._add_alternating_row_colors(ws)

        # Add borders
        self._add_borders(ws)

        # Format cost column
        self._format_cost_column(ws)

        # Restore title
        ws['A3'] = title_cell

        wb.save(filename)

    def _style_headers(self, ws):
        """Style the header rows."""
        row2 = ws[1]
        for col_idx, cell in enumerate(row2, 1):
            for section, (start, end) in self.sections.items():
                if start <= col_idx <= end:
                    subheader_color = self.section_colors[section]['subheader'].replace('#', '')
                    cell.fill = PatternFill(start_color=subheader_color,
                                            end_color=subheader_color,
                                            fill_type='solid')
                    cell.font = Font(bold=True, color='444444')
                    cell.alignment = Alignment(horizontal='center',
                                               vertical='center',
                                               wrap_text=True)
                    break

    def _add_sections(self, ws):
        """Add and style section headers."""
        ws.insert_rows(1)
        current_col = 1
        for section, (start, end) in self.sections.items():
            if current_col <= end:
                cell = ws.cell(row=1, column=current_col)
                cell.value = section
                cell.font = Font(bold=True, color='FFFFFF')
                header_color = self.section_colors[section]['header'].replace('#', '')
                cell.fill = PatternFill(start_color=header_color,
                                        end_color=header_color,
                                        fill_type='solid')
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center',
                                           wrap_text=True)

                if start != end:
                    ws.merge_cells(start_row=1, start_column=current_col,
                                   end_row=1, end_column=end)
                current_col = end + 1

    def _format_data_cells(self, ws):
        """Format data cells including special URL formatting."""
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                if cell.column == 2:  # URL column
                    cell.alignment = Alignment(horizontal='left',
                                               vertical='center',
                                               wrap_text=False)
                    cell.font = Font(color='0563C1', underline='single')
                    if cell.value and isinstance(cell.value, str):
                        cell.hyperlink = cell.value
                else:
                    cell.alignment = Alignment(horizontal='left',
                                               vertical='center',
                                               wrap_text=True)
                    cell.font = Font(color='444444')

    def _add_alternating_row_colors(self, ws):
        """Add alternating row colors for better readability."""
        for row in range(3, ws.max_row + 1):
            fill_color = 'F7F9FB' if row % 2 == 0 else 'FFFFFF'
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if not cell.fill.start_color.rgb:
                    cell.fill = PatternFill(start_color=fill_color,
                                            end_color=fill_color,
                                            fill_type='solid')

    def _add_borders(self, ws):
        """Add thin borders to all cells."""
        thin_border = Border(
            left=Side(style='thin', color='E3E3E3'),
            right=Side(style='thin', color='E3E3E3'),
            top=Side(style='thin', color='E3E3E3'),
            bottom=Side(style='thin', color='E3E3E3')
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

    def _format_cost_column(self, ws):
        """Format the API cost column."""
        header_row = [cell.value for cell in ws[2]]
        cost_col = header_row.index('API Cost') + 1
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=cost_col)
            cell.number_format = '$#,##0.00000'


def style_excel_file(filename):
    """Convenience function to style an Excel file."""
    styler = ExcelStyler()
    styler.style_excel_file(filename)


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 2:
        print("Usage: python excel_styler.py <excel_file>")
        sys.exit(1)

    filename = sys.argv[1]
    style_excel_file(filename)