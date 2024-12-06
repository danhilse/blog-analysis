import json
from textstat import textstat
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import pandas as pd
from datetime import datetime
from urllib.parse import urlparse
import os
from typing import Dict, List, Tuple
import argparse
import sys

# Ensure that 'ai_analysis' module and its functions are correctly imported
# from ai_analysis import analyze_content_categorization, cost_tracker, analyze_seo, analyze_tone_voice, analyze_quality_brand_fit


def add_no_match_highlighting(ws):
    # (Your existing add_no_match_highlighting function)
    header_row = [cell.value for cell in ws[2]]

    target_columns = [
        'Primary Category',
        'Solution Topic',
        'Use Case',
        'Customer Journey Stage',
        'CMO Priority',
        'Marketing Activity Type',
        'Target Audience'
    ]

    column_indices = {col: header_row.index(col) + 1 for col in target_columns if col in header_row}

    highlight_values = [
        'No Clear Match',
        'NONE',
        'No Clear Topic',
        'No Clear Activity Type',
        'No Clear Audience'
    ]

    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)

    for row in range(3, ws.max_row + 1):
        for col_name, col_idx in column_indices.items():
            cell = ws.cell(row=row, column=col_idx)
            if cell.value in highlight_values:
                cell.fill = red_fill
                cell.font = white_font

def add_conditional_formatting(ws):
    # (Your existing add_conditional_formatting function)
    header_row = [cell.value for cell in ws[2]]

    columns = {
        'Overall Quality Score': header_row.index('Overall Quality Score') + 1,
        'Topic Relevance': header_row.index('Topic Relevance') + 1,
        'Brand Alignment': header_row.index('Brand Alignment') + 1,
        'Natural/Conversational Score': header_row.index('Natural/Conversational Score') + 1,
        'Authentic/Approachable Score': header_row.index('Authentic/Approachable Score') + 1,
        'Gender-Neutral/Inclusive Score': header_row.index('Gender-Neutral/Inclusive Score') + 1,
        'Personal Pronoun Count': header_row.index('Personal Pronoun Count') + 1,
        'Keyword Integration Score': header_row.index('Keyword Integration Score') + 1,
        'Meta Description Quality Score': header_row.index('Meta Description Quality Score') + 1,
        'Outdated Widgets Count': header_row.index('Outdated Widgets Count') + 1,
        'Word Count': header_row.index('Word Count') + 1
    }

    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow_green_fill = PatternFill(start_color='9ACD32', end_color='9ACD32', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    score_columns = ['Overall Quality Score', 'Natural/Conversational Score',
                     'Authentic/Approachable Score', 'Gender-Neutral/Inclusive Score',
                     'Keyword Integration Score', 'Meta Description Quality Score']

    for col_name in score_columns:
        col_letter = get_column_letter(columns[col_name])
        ws.conditional_formatting.add(
            f'{col_letter}3:{col_letter}{ws.max_row}',
            ColorScaleRule(
                start_type='num', start_value=0, start_color='FF0000',
                mid_type='num', mid_value=50, mid_color='FFFF00',
                end_type='num', end_value=100, end_color='00FF00'
            )
        )

    word_count_col = get_column_letter(columns['Word Count'])
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=columns['Word Count'])
        try:
            count = int(cell.value)
            if 1000 <= count <= 1200:
                cell.fill = green_fill
            elif (800 <= count < 1000) or (1201 <= count <= 1400):
                cell.fill = yellow_fill
            else:
                cell.fill = orange_fill
        except (ValueError, TypeError):
            pass

    topic_col = get_column_letter(columns['Topic Relevance'])
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=columns['Topic Relevance'])
        if cell.value == "Tangentially Related":
            cell.fill = yellow_fill
        elif cell.value == "Off Topic":
            cell.fill = red_fill

    brand_col = get_column_letter(columns['Brand Alignment'])
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=columns['Brand Alignment'])
        if cell.value == "On Brand":
            cell.fill = green_fill
        elif cell.value == "Mostly on Brand":
            cell.fill = yellow_green_fill
        elif cell.value == "Needs Work":
            cell.fill = yellow_fill
        elif cell.value == "Not on Brand":
            cell.fill = red_fill

    widgets_col = get_column_letter(columns['Outdated Widgets Count'])
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=widgets_col)
        try:
            if int(cell.value) > 0:
                cell.fill = red_fill
        except (ValueError, TypeError):
            pass

    pronoun_col = get_column_letter(columns['Personal Pronoun Count'])
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=pronoun_col)
        try:
            if int(cell.value) > 0:
                cell.fill = red_fill
        except (ValueError, TypeError):
            pass

def style_excel_file(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Check if headers already exist by inspecting a specific cell
    if ws['A1'].value != 'Title':
        # Extract title from first row and remove it
        title_cell = ws['A2'].value  # Save title value

        # Define brand colors with lighter shades for subheaders
        section_colors = {
            'Title': {
                'header': '#193661',  # Dark blue
                'subheader': '#C1C9D4'  # Lighter blue
            },
            'Basic Information': {
                'header': '#00babe',  # Teal
                'subheader': '#E5F9F9'  # Lighter teal
            },
            'Quality & Brand Fit': {
                'header': '#e34e64',  # Salmon
                'subheader': '#FFEFEF'  # Lighter salmon
            },
            'Tone & Voice': {
                'header': '#193661',  # Dark blue
                'subheader': '#C1C9D4'  # Lighter blue
            },
            'SEO Analysis': {
                'header': '#00babe',  # Teal
                'subheader': '#E5F9F9'  # Lighter teal
            },
            'Multimedia Assessment': {
                'header': '#e34e64',  # Salmon
                'subheader': '#FFEFEF'  # Lighter salmon
            },
            'Content Categorization': {
                'header': '#193661',  # Dark blue
                'subheader': '#C1C9D4'  # Lighter blue
            },
            'Performance Metrics': {
                'header': '#00babe',  # Teal
                'subheader': '#E5F9F9'  # Lighter teal
            },
            'Cost Analysis': {            # Added missing section
                'header': '#e34e64',      # Salmon
                'subheader': '#FFEFEF'    # Lighter salmon
            }
        }

        # Define sections and their ranges
        sections = {
            'Title': (1, 1),
            'Basic Information': (2, 7),  # Updated to include Modified Date
            'Quality & Brand Fit': (8, 12),
            'Tone & Voice': (13, 19),
            'SEO Analysis': (20, 29),  # Updated for split H2/H3
            'Multimedia Assessment': (30, 36),
            'Content Categorization': (37, 43),
            'Performance Metrics': (44, 49),
            'Cost Analysis': (50, 50)  # New section
        }

        # Freeze top two rows and first column
        ws.freeze_panes = 'B3'

        # Set specific column widths
        ws.column_dimensions['B'].width = 15  # URL column
        for col in range(1, ws.max_column + 1):
            if get_column_letter(col) != 'B':  # All non-URL columns
                ws.column_dimensions[get_column_letter(col)].width = 30

        # Style metric headers (second row)
        row2 = ws[2]  # Get the second row (index 2)
        for col_idx, cell in enumerate(row2, 1):
            # Find which section this column belongs to
            for section, (start, end) in sections.items():
                if start <= col_idx <= end:
                    subheader_color = section_colors[section]['subheader'].replace('#', '')
                    cell.fill = PatternFill(start_color=subheader_color,
                                          end_color=subheader_color,
                                          fill_type='solid')
                    cell.font = Font(bold=True, color='444444')
                    cell.alignment = Alignment(horizontal='center',
                                             vertical='center',
                                             wrap_text=True)
                    break

        # Add and style section headers
        ws.insert_rows(1)
        current_col = 1
        for section, (start, end) in sections.items():
            if current_col <= end:
                cell = ws.cell(row=1, column=current_col)
                cell.value = section
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=section_colors[section]['header'].replace('#', ''),
                                      end_color=section_colors[section]['header'].replace('#', ''),
                                      fill_type='solid')
                cell.alignment = Alignment(horizontal='center',
                                         vertical='center',
                                         wrap_text=True)

                if start != end:
                    ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=end)
                current_col = end + 1

        # Get column indices for conditional formatting
        header_row = [cell.value for cell in ws[2]]  # Get header row values
        try:
            reading_level_col = header_row.index('Reading Level (Gunning Fog)') + 1
            header_image_width_col = header_row.index('Header Image Width') + 1
        except ValueError:
            reading_level_col = None
            header_image_width_col = None

        # Add conditional formatting for Word Count and Reading Level
        if reading_level_col:
            ws.conditional_formatting.add(
                f'{get_column_letter(reading_level_col)}3:{get_column_letter(reading_level_col)}{ws.max_row}',
                ColorScaleRule(
                    start_type='min',
                    start_color='FF0000',  # Red
                    mid_type='percentile',
                    mid_value=50,
                    mid_color='FFFF00',  # Yellow
                    end_type='max',
                    end_color='00FF00'  # Green
                )
            )

        # Add conditional formatting for Header Image Width
        if header_image_width_col:
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=header_image_width_col)
                try:
                    width = int(cell.value)
                    if width >= 1200:
                        cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
                    elif width >= 800:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
                    else:
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
                except (ValueError, TypeError):
                    pass

        # Format data cells (including URL special formatting)
        for row in ws.iter_rows(min_row=3):  # Start from data rows
            for cell in row:
                # Special formatting for URL column
                if cell.column == 2:  # URL column (B)
                    cell.alignment = Alignment(horizontal='left',
                                             vertical='center',
                                             wrap_text=False)
                    cell.font = Font(color='0563C1', underline='single')
                    if cell.value and isinstance(cell.value, str):
                        cell.hyperlink = cell.value
                else:
                    cell.alignment = Alignment(horizontal='left',
                                             vertical='center',
                                             wrap_text=True)
                    cell.font = Font(color='444444')

        # Add alternating row colors for better readability
        for row in range(3, ws.max_row + 1):
            fill_color = 'F7F9FB' if row % 2 == 0 else 'FFFFFF'
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if not cell.fill.start_color.rgb:  # Only apply if no other fill color is present
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        # Add thin borders to all cells
        thin_border = Border(
            left=Side(style='thin', color='E3E3E3'),
            right=Side(style='thin', color='E3E3E3'),
            top=Side(style='thin', color='E3E3E3'),
            bottom=Side(style='thin', color='E3E3E3')
        )

        # Add the new highlighting for "No Clear Match" values
        add_no_match_highlighting(ws)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        try:
            cost_col = header_row.index('API Cost') + 1
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=cost_col)
                cell.number_format = '$#,##0.00000'
        except ValueError:
            pass

        add_conditional_formatting(ws)

        # Restore title in its new position
        if ws['A3'].value is None and 'title_cell' in locals():
            ws['A3'] = title_cell

        wb.save(filename)

def create_blog_audit_df(articles_data, start_index=0):
    """
    Creates a list of dictionaries containing audit information for multiple articles.
    Includes performance metrics from Google Analytics data.

    Parameters:
        articles_data (list): List of articles to process.
        start_index (int): The starting index of the batch in the entire dataset.

    Returns:
        list: List of dictionaries containing audit information.
    """
    total_articles = len(articles_data)
    url_to_keyword = load_yoast_keywords()
    all_articles = []
    performance_df = import_performance_data()

    last_successful_index = start_index - 1

    for idx, article in enumerate(articles_data):
        i = start_index + idx  # Actual index in the full list
        cost_tracker.reset()  # Reset cost tracker for new article

        try:
            # Safely extract 'content'
            content = article.get('content', '')
            clean_text = clean_content(content)
            print(f"\nArticle {i}/{total_articles + start_index}")

            # Safely extract 'basic_info'
            basic_info = article.get('basic_info', {})
            url = basic_info.get('url', 'No URL')
            title = basic_info.get('title', 'No Title')
            publication_date_str = basic_info.get('publication_date')
            modified_date_str = basic_info.get('modified_date')

            # Get pronoun analysis
            pronoun_analysis = count_personal_pronouns(content)

            # Extract target keyword
            target_keyword = url_to_keyword.get(url, 'Not Found')

            # Extract slug for performance data matching
            slug = urlparse(url).path.strip('/').split('/')[-1] if url != 'No URL' else 'No Slug'

            # Safely extract 'multimedia_assessment'
            multimedia = article.get('multimedia_assessment', {})
            header_image = multimedia.get('header_image') or {}

            # Get outdated widget count from multimedia assessment
            outdated_widgets_count = multimedia.get('outdated_widget_count', 0)


            # Safeguard against missing 'width' and 'height'
            try:
                header_width = int(header_image.get('width', 0))
            except (TypeError, ValueError):
                header_width = 0

            try:
                header_height = int(header_image.get('height', 0))
            except (TypeError, ValueError):
                header_height = 0

            header_src = header_image.get('src', 'No Source')
            header_alt = header_image.get('alt', 'No Alt Text')

            # Safely extract 'content_images'
            content_images = multimedia.get('content_images', [])
            content_widths = []
            for img in content_images:
                try:
                    width = int(img.get('width', 0))
                    content_widths.append(width)
                except (TypeError, ValueError):
                    content_widths.append(0)

            min_content_width = min(content_widths) if content_widths else 0

            # Get performance metrics if available
            performance_metrics = {
                'Total Views': 0,
                'Total Users': 0,
                'Total Sessions': 0,
                'Engagement Rate': 0.0,
                'Average Time on Page': 0.0,
                'Bounce Rate': 0.0
            }

            if performance_df is not None and slug in performance_df.index:
                metrics = performance_df.loc[slug]
                performance_metrics = {
                    'Total Views': metrics.get('Views', 0),
                    'Total Users': metrics.get('Total users', 0),
                    'Total Sessions': metrics.get('Sessions', 0),
                    'Engagement Rate': metrics.get('Engagement rate', 0.0),
                    'Average Time on Page': metrics.get('Average session duration', 0.0),
                    'Bounce Rate': metrics.get('Bounce rate', 0.0)
                }

            # Safely extract 'seo_analysis'
            seo_analysis = article.get('seo_analysis', {})
            meta_description = seo_analysis.get('meta_description', {})
            meta_present = meta_description.get('present', False)

            headings = seo_analysis.get('headings', {})
            h1_present = headings.get('h1_present', False)
            h2_count = headings.get('h2_count', 0)
            h3_count = headings.get('h3_count', 0)

            # Parse dates safely
            def parse_date(date_str):
                if date_str:
                    try:
                        return datetime.strptime(date_str.split('T')[0], '%Y-%m-%d').strftime('%Y-%m-%d')
                    except ValueError:
                        return 'No Date'
                return 'No Date'

            publication_date = parse_date(publication_date_str)
            modified_date = parse_date(modified_date_str)

            print("  ↳ Content categorization...")
            categorization = analyze_content_categorization(clean_text)

            # Format and analyze SEO
            print("  ↳ SEO analysis...")
            seo_data = format_seo_data(basic_info, seo_analysis, target_keyword)
            seo_results = analyze_seo(clean_text, seo_data)

            # Get tone and voice analysis
            print("  ↳ Tone and voice analysis...")
            tone_results = analyze_tone_voice(clean_text)

            # Get quality and brand fit analysis
            print("  ↳ Quality and brand analysis...")
            quality_results = analyze_quality_brand_fit(clean_text)

            # Extract SEO results with error handling
            try:
                keyword_density = float(seo_results.get("Keyword Density", 0))
                keyword_integration = int(seo_results.get("Keyword Integration Score", 0))
                meta_quality = int(seo_results.get("Meta Description Quality Score", 0))
                recommended_keywords = seo_results.get("Recommended New Keywords", [])
                seo_notes = seo_results.get("SEO Notes/Recommendations", "No recommendations available")
            except (ValueError, TypeError, AttributeError) as e:
                print(f"Error processing SEO results for {url}: {str(e)}")
                keyword_density = 0
                keyword_integration = 0
                meta_quality = 0
                recommended_keywords = []
                seo_notes = "Error processing SEO analysis"

            # Extract tone/voice results with error handling
            try:
                challenger_pct = int(tone_results.get("Challenger Percentage", 0))
                supportive_pct = int(tone_results.get("Supportive Percentage", 0))
                natural_score = int(tone_results.get("Natural/Conversational Score", 0))
                authentic_score = int(tone_results.get("Authentic/Approachable Score", 0))
                inclusive_score = int(tone_results.get("Gender-Neutral/Inclusive Score", 0))
                tone_notes = tone_results.get("Tone Notes/Recommendations", "No recommendations available")
            except (ValueError, TypeError, AttributeError) as e:
                print(f"Error processing tone results for {url}: {str(e)}")
                challenger_pct = 0
                supportive_pct = 0
                natural_score = 0
                authentic_score = 0
                inclusive_score = 0
                tone_notes = "Error processing tone analysis"

            # Extract quality/brand results with error handling
            try:
                quality_results = analyze_quality_brand_fit(clean_text)
                if quality_results:  # Check if we got a response
                    quality_score = int(quality_results.get("Overall Quality Score", 0))
                    brand_alignment = quality_results.get("Brand Alignment", "Needs Work")
                    quality_notes = quality_results.get("Quality Notes", "No recommendations available")
                    brand_alignment_notes = quality_results.get("Brand Alignment Notes",
                                                                "No brand alignment notes available")
                else:
                    raise ValueError("No quality results returned")
            except (ValueError, TypeError, AttributeError) as e:
                print(f"Error processing quality results for {url}: {str(e)}")
                quality_score = 0
                brand_alignment = "Error"
                quality_notes = "Error processing quality analysis"
                brand_alignment_notes = "Error processing brand alignment analysis"

            article_data = {
                # Basic Information
                'Title': title,
                'URL': url,
                'Slug': slug,
                'Publication Date': publication_date,
                'Modified Date': modified_date,
                'Word Count': calculate_word_count(clean_text),
                'Reading Level (Gunning Fog)': round(textstat.gunning_fog(clean_text), 1) if clean_text else 0.0,

                # Quality & Brand Fit
                'Overall Quality Score': quality_score,
                'Topic Relevance': quality_results.get("Topic Relevance", "Error"),
                'Brand Alignment': brand_alignment,
                'Quality Notes/Recommendations': quality_notes,
                'Brand Alignment Notes': brand_alignment_notes,

                # Tone & Voice
                'Challenger Percentage': f"{challenger_pct}%",
                'Supportive Percentage': f"{supportive_pct}%",
                'Natural/Conversational Score': natural_score,
                'Authentic/Approachable Score': authentic_score,
                'Gender-Neutral/Inclusive Score': inclusive_score,
                'Tone Notes/Recommendations': tone_notes,
                'Personal Pronoun Count': pronoun_analysis['count'],

                # SEO Analysis
                'Current Target Keyword': target_keyword,
                'Keyword Density': f"{keyword_density:.2f}%",
                'Keyword Integration Score': keyword_integration,
                'Meta Description Present': 'Yes' if meta_present else 'No',
                'Meta Description Quality Score': meta_quality,
                'H1 Tag Present': 'Yes' if h1_present else 'No',
                'H2 Tags': h2_count,
                'H3 Tags': h3_count,
                'Recommended New Keywords': ' | '.join(recommended_keywords) if recommended_keywords else 'None',
                'SEO Notes/Recommendations': seo_notes,

                # Multimedia Assessment
                'Number of Images': multimedia.get('total_image_count', 0),
                'Header Image Width': header_width,
                'Header Image Height': header_height,
                'Header Image Src': header_src,
                'Header Image Alt': header_alt,
                'Minimum Content Image Width': min_content_width,
                'Outdated Widgets Count': outdated_widgets_count,

                # Content Categorization
                'Primary Category': categorization["Primary Category"],
                'Solution Topic': categorization["Solution Topic"],
                'Use Case': categorization["Use Case"],
                'Customer Journey Stage': categorization["Customer Journey Stage"],
                'CMO Priority': categorization["CMO Priority"],
                'Marketing Activity Type': categorization["Marketing Activity Type"],
                'Target Audience': categorization["Target Audience"],

                # Performance Metrics (from Google Analytics)
                **performance_metrics,

                'API Cost': f"${cost_tracker.cost:.4f}",  # Add cost as last field
            }

            print(f"  ✓ Complete (Cost: ${cost_tracker.cost:.4f})")

            all_articles.append(article_data)
            last_successful_index = i

        except Exception as e:
            print(f"Error processing article at index {i}: {str(e)}")
            print(f"Last successfully processed article index: {last_successful_index}")
            break  # Stop processing further articles

        if not all_articles:
            print("No articles were successfully processed.")
            return []

        return all_articles

def load_yoast_keywords():
    """Load Yoast keywords data and create URL to keyword mapping"""
    try:
        # Read the Yoast keywords Excel file
        yoast_df = pd.read_excel('resources/yoast-blog-keywords.xlsx', header=None)

        # Create a dictionary mapping URLs to keywords
        url_to_keyword = {}
        for _, row in yoast_df.iterrows():
            url = row[4]  # URL is in column 5 (index 4)
            keyword = row[2]  # Keyword is in column 3 (index 2)
            if pd.notna(url) and pd.notna(keyword):
                url_to_keyword[url.strip()] = keyword.strip()

        return url_to_keyword
    except Exception as e:
        print(f"Error loading Yoast keywords: {e}")
        return {}

def import_performance_data():
    """
    Imports performance data from resources/performance.xlsx and processes it for integration
    with blog audit data.
    """
    try:
        file_path = os.path.join('resources', 'performance.xlsx')
        performance_df = pd.read_excel(file_path)

        def extract_slug(path):
            if pd.isna(path):
                return ''
            parsed = urlparse(path)
            path_parts = [p for p in parsed.path.strip('/').split('/') if p]
            return path_parts[-1] if path_parts else ''

        performance_df['slug'] = performance_df['Page path + query string'].apply(extract_slug)
        performance_df = performance_df[performance_df['slug'] != '']
        performance_df.set_index('slug', inplace=True)

        return performance_df
    except Exception as e:
        print(f"Warning: Could not load performance data: {str(e)}")
        return None

def count_personal_pronouns(text: str) -> Dict[str, any]:
    """
    Counts personal pronouns while excluding those within quotes.
    Additionally, prints the sentences that contain pronouns outside quotes.
    Handles straight (") and curly (“ and ”) double quotes.
    """
    # Define quote pairs: opening quote -> closing quote
    quote_pairs = {
        '"': '"',
        '“': '”',
        '‘': '’',  # Include if handling single curly quotes
    }

    quotes: List[Tuple[int, int]] = []
    current_pos = 0
    text_length = len(text)

    # Step 1: Detect quoted regions
    while current_pos < text_length:
        # Find the next opening quote
        next_quote = None
        next_pos = text_length

        for open_quote in quote_pairs.keys():
            pos = text.find(open_quote, current_pos)
            if pos != -1 and pos < next_pos:
                next_quote = open_quote
                next_pos = pos

        if next_quote is None:
            break  # No more quotes found

        start = next_pos
        end_quote = quote_pairs[next_quote]
        end = text.find(end_quote, start + 1)

        if end == -1:
            # If no closing quote is found, consider the rest of the text as quoted
            end = text_length - 1

        quotes.append((start, end))
        current_pos = end + 1

    def is_in_quotes(pos: int) -> bool:
        """Check if a position is within any quoted region."""
        return any(start <= pos <= end for start, end in quotes)

    # Step 2: Split text into sentences with their positions
    sentence_endings = re.compile(r'([.!?])')  # Regex to identify sentence boundaries
    sentences: List[Tuple[str, int, int]] = []
    start = 0

    for match in sentence_endings.finditer(text):
        end = match.end()
        sentence = text[start:end].strip()
        if sentence:
            sentences.append((sentence, start, end))
        start = end

    # Add any remaining text as the last sentence
    if start < text_length:
        sentence = text[start:].strip()
        if sentence:
            sentences.append((sentence, start, text_length))

    # Step 3: Find pronouns outside quotes
    pronouns = r'\b(I|me|my|mine|myself)\b'
    matches = []
    sentences_with_pronouns: List[str] = []
    seen_sentences: set = set()

    for match in re.finditer(pronouns, text, re.IGNORECASE):
        if not is_in_quotes(match.start()):
            matches.append(match)
            # Step 4: Find the sentence containing this pronoun
            for sentence, s_start, s_end in sentences:
                if s_start <= match.start() < s_end:
                    if sentence not in seen_sentences:
                        sentences_with_pronouns.append(sentence)
                        seen_sentences.add(sentence)
                    break

    return {
        'count': len(matches),
        'found_pronouns': [m.group() for m in matches],
        'quoted_regions': quotes,
        'sentences_with_pronouns': sentences_with_pronouns,
        'flag': len(matches) > 0
    }

def write_headers(ws):
    """
    Write the two header rows to the worksheet.

    Parameters:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to write headers to.
    """
    # Define your two header rows
    headers_row1 = [
        'Title', 'Basic Information', 'Quality & Brand Fit', 'Tone & Voice',
        'SEO Analysis', 'Multimedia Assessment', 'Content Categorization',
        'Performance Metrics', 'Cost Analysis'
    ]

    # Define corresponding second row headers
    headers_row2 = [
        'Title', 'URL', 'Slug', 'Publication Date', 'Modified Date', 'Word Count',
        'Reading Level (Gunning Fog)', 'Overall Quality Score', 'Topic Relevance',
        'Brand Alignment', 'Quality Notes/Recommendations', 'Brand Alignment Notes',
        'Challenger Percentage', 'Supportive Percentage', 'Natural/Conversational Score',
        'Authentic/Approachable Score', 'Gender-Neutral/Inclusive Score',
        'Tone Notes/Recommendations', 'Personal Pronoun Count',
        'Current Target Keyword', 'Keyword Density', 'Keyword Integration Score',
        'Meta Description Present', 'Meta Description Quality Score', 'H1 Tag Present',
        'H2 Tags', 'H3 Tags', 'Recommended New Keywords', 'SEO Notes/Recommendations',
        'Number of Images', 'Header Image Width', 'Header Image Height', 'Header Image Src',
        'Header Image Alt', 'Minimum Content Image Width', 'Outdated Widgets Count',
        'Primary Category', 'Solution Topic', 'Use Case', 'Customer Journey Stage',
        'CMO Priority', 'Marketing Activity Type', 'Target Audience',
        'Total Views', 'Total Users', 'Total Sessions', 'Engagement Rate',
        'Average Time on Page', 'Bounce Rate', 'API Cost'
    ]

    ws.append(headers_row1)
    ws.append(headers_row2)

def append_data(ws, data: List[Dict]):
    """
    Append data rows to the worksheet.

    Parameters:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to append data to.
        data (list of dict): List of dictionaries representing data rows.
    """
    for row in data:
        row_values = [
            row.get('Title', ''),
            row.get('URL', ''),
            row.get('Slug', ''),
            row.get('Publication Date', ''),
            row.get('Modified Date', ''),
            row.get('Word Count', 0),
            row.get('Reading Level (Gunning Fog)', 0.0),
            row.get('Overall Quality Score', 0),
            row.get('Topic Relevance', ''),
            row.get('Brand Alignment', ''),
            row.get('Quality Notes/Recommendations', ''),
            row.get('Brand Alignment Notes', ''),
            row.get('Challenger Percentage', ''),
            row.get('Supportive Percentage', ''),
            row.get('Natural/Conversational Score', 0),
            row.get('Authentic/Approachable Score', 0),
            row.get('Gender-Neutral/Inclusive Score', 0),
            row.get('Tone Notes/Recommendations', ''),
            row.get('Personal Pronoun Count', 0),
            row.get('Current Target Keyword', ''),
            row.get('Keyword Density', '0.00%'),
            row.get('Keyword Integration Score', 0),
            row.get('Meta Description Present', ''),
            row.get('Meta Description Quality Score', 0),
            row.get('H1 Tag Present', ''),
            row.get('H2 Tags', 0),
            row.get('H3 Tags', 0),
            row.get('Recommended New Keywords', 'None'),
            row.get('SEO Notes/Recommendations', ''),
            row.get('Number of Images', 0),
            row.get('Header Image Width', 0),
            row.get('Header Image Height', 0),
            row.get('Header Image Src', ''),
            row.get('Header Image Alt', ''),
            row.get('Minimum Content Image Width', 0),
            row.get('Outdated Widgets Count', 0),
            row.get('Primary Category', ''),
            row.get('Solution Topic', ''),
            row.get('Use Case', ''),
            row.get('Customer Journey Stage', ''),
            row.get('CMO Priority', ''),
            row.get('Marketing Activity Type', ''),
            row.get('Target Audience', ''),
            row.get('Total Views', 0),
            row.get('Total Users', 0),
            row.get('Total Sessions', 0),
            row.get('Engagement Rate', 0.0),
            row.get('Average Time on Page', 0.0),
            row.get('Bounce Rate', 0.0),
            row.get('API Cost', '$0.0000')
        ]
        ws.append(row_values)

def style_excel_file(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Check if headers already exist by inspecting a specific cell
    if ws['A1'].value != 'Title':
        # If headers do not exist, write them
        write_headers(ws)

    # Define brand colors with lighter shades for subheaders
    section_colors = {
        'Title': {
            'header': '#193661',  # Dark blue
            'subheader': '#C1C9D4'  # Lighter blue
        },
        'Basic Information': {
            'header': '#00babe',  # Teal
            'subheader': '#E5F9F9'  # Lighter teal
        },
        'Quality & Brand Fit': {
            'header': '#e34e64',  # Salmon
            'subheader': '#FFEFEF'  # Lighter salmon
        },
        'Tone & Voice': {
            'header': '#193661',  # Dark blue
            'subheader': '#C1C9D4'  # Lighter blue
        },
        'SEO Analysis': {
            'header': '#00babe',  # Teal
            'subheader': '#E5F9F9'  # Lighter teal
        },
        'Multimedia Assessment': {
            'header': '#e34e64',  # Salmon
            'subheader': '#FFEFEF'  # Lighter salmon
        },
        'Content Categorization': {
            'header': '#193661',  # Dark blue
            'subheader': '#C1C9D4'  # Lighter blue
        },
        'Performance Metrics': {
            'header': '#00babe',  # Teal
            'subheader': '#E5F9F9'  # Lighter teal
        },
        'Cost Analysis': {            # Added missing section
            'header': '#e34e64',      # Salmon
            'subheader': '#FFEFEF'    # Lighter salmon
        }
    }

    # Define sections and their ranges
    sections = {
        'Title': (1, 1),
        'Basic Information': (2, 7),  # Updated to include Modified Date
        'Quality & Brand Fit': (8, 12),
        'Tone & Voice': (13, 19),
        'SEO Analysis': (20, 29),  # Updated for split H2/H3
        'Multimedia Assessment': (30, 36),
        'Content Categorization': (37, 43),
        'Performance Metrics': (44, 49),
        'Cost Analysis': (50, 50)  # New section
    }

    # Freeze top two rows and first column
    ws.freeze_panes = 'B3'

    # Set specific column widths
    ws.column_dimensions['B'].width = 15  # URL column
    for col in range(1, ws.max_column + 1):
        if get_column_letter(col) != 'B':  # All non-URL columns
            ws.column_dimensions[get_column_letter(col)].width = 30

    # Style metric headers (second row)
    row2 = ws[2]  # Get the second row (index 2)
    for col_idx, cell in enumerate(row2, 1):
        # Find which section this column belongs to
        for section, (start, end) in sections.items():
            if start <= col_idx <= end:
                subheader_color = section_colors[section]['subheader'].replace('#', '')
                cell.fill = PatternFill(start_color=subheader_color,
                                      end_color=subheader_color,
                                      fill_type='solid')
                cell.font = Font(bold=True, color='444444')
                cell.alignment = Alignment(horizontal='center',
                                         vertical='center',
                                         wrap_text=True)
                break

    # Add and style section headers
    # Assuming headers are already written, skip adding them again

    # Get column indices for conditional formatting
    header_row = [cell.value for cell in ws[2]]  # Get header row values
    try:
        reading_level_col = header_row.index('Reading Level (Gunning Fog)') + 1
        header_image_width_col = header_row.index('Header Image Width') + 1
    except ValueError:
        reading_level_col = None
        header_image_width_col = None

    # Add conditional formatting for Word Count and Reading Level
    if reading_level_col:
        ws.conditional_formatting.add(
            f'{get_column_letter(reading_level_col)}3:{get_column_letter(reading_level_col)}{ws.max_row}',
            ColorScaleRule(
                start_type='min',
                start_color='FF0000',  # Red
                mid_type='percentile',
                mid_value=50,
                mid_color='FFFF00',  # Yellow
                end_type='max',
                end_color='00FF00'  # Green
            )
        )

    # Add conditional formatting for Header Image Width
    if header_image_width_col:
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=header_image_width_col)
            try:
                width = int(cell.value)
                if width >= 1200:
                    cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
                elif width >= 800:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
                else:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
            except (ValueError, TypeError):
                pass

    # Format data cells (including URL special formatting)
    for row in ws.iter_rows(min_row=3):  # Start from data rows
        for cell in row:
            # Special formatting for URL column
            if cell.column == 2:  # URL column (B)
                cell.alignment = Alignment(horizontal='left',
                                         vertical='center',
                                         wrap_text=False)
                cell.font = Font(color='0563C1', underline='single')
                if cell.value and isinstance(cell.value, str):
                    cell.hyperlink = cell.value
            else:
                cell.alignment = Alignment(horizontal='left',
                                         vertical='center',
                                         wrap_text=True)
                cell.font = Font(color='444444')

    # Add alternating row colors for better readability
    for row in range(3, ws.max_row + 1):
        fill_color = 'F7F9FB' if row % 2 == 0 else 'FFFFFF'
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if not cell.fill.start_color.rgb:  # Only apply if no other fill color is present
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    # Add thin borders to all cells
    thin_border = Border(
        left=Side(style='thin', color='E3E3E3'),
        right=Side(style='thin', color='E3E3E3'),
        top=Side(style='thin', color='E3E3E3'),
        bottom=Side(style='thin', color='E3E3E3')
    )

    # Add the new highlighting for "No Clear Match" values
    add_no_match_highlighting(ws)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    try:
        cost_col = header_row.index('API Cost') + 1
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=cost_col)
            cell.number_format = '$#,##0.00000'
    except ValueError:
        pass

    add_conditional_formatting(ws)

    # Restore title in its new position
    if ws['A3'].value is None:
        ws['A3'] = 'Title'  # Replace with actual title if needed

    wb.save(filename)

def main():
    parser = argparse.ArgumentParser(description='Process articles in batches.')
    parser.add_argument('--start_index', type=int, default=0, help='Index to start processing from.')
    parser.add_argument('--batch_size', type=int, default=100, help='Number of articles to process in this batch.')

    args = parser.parse_args()

    input_file = 'output/all.json'
    output_file = 'output/blog_audit.xlsx'

    # Read the JSON file
    try:
        with open(input_file, 'r') as file:
            data = json.load(file)
    except Exception as e:
        print(f"Failed to read input JSON file: {e}")
        sys.exit(1)

    # Assuming 'data' has a structure like {'analyses': {'blog': [...]}}
    articles = data.get('analyses', {}).get('blog', [])
    total_articles = len(articles)

    if total_articles == 0:
        print("No articles found in the JSON data.")
        sys.exit(0)

    # Calculate the end index
    start_index = args.start_index
    batch_size = args.batch_size
    end_index = min(start_index + batch_size, total_articles)

    # Get the batch of articles
    articles_batch = articles[start_index:end_index]

    if not articles_batch:
        print("No articles to process in this batch.")
        sys.exit(0)

    # Process the data
    try:
        data_rows = create_blog_audit_df(articles_batch, start_index=start_index)
    except Exception as e:
        print(f"An error occurred while processing the batch: {e}")
        sys.exit(1)

    if not data_rows:
        print("No new articles were processed.")
        sys.exit(0)

    # Check if the Excel file exists
    file_exists = os.path.exists(output_file)

    if not file_exists:
        # Create a new workbook and add headers
        wb = openpyxl.Workbook()
        ws = wb.active
        write_headers(ws)
        # Append data rows
        append_data(ws, data_rows)
        # Save the workbook
        wb.save(output_file)
        # Apply styling
        style_excel_file(output_file)
        print(f"Batch processed: Articles {start_index} to {end_index - 1}")
    else:
        # Open the existing workbook
        try:
            wb = openpyxl.load_workbook(output_file)
            ws = wb.active
        except Exception as e:
            print(f"Failed to open existing Excel file: {e}")
            sys.exit(1)

        # Append data rows
        append_data(ws, data_rows)

        # Save the workbook
        try:
            wb.save(output_file)
        except Exception as e:
            print(f"Failed to save Excel file: {e}")
            sys.exit(1)

        print(f"Batch processed: Articles {start_index} to {end_index - 1}")

if __name__ == "__main__":
    main()